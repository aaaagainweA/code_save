#自动化播报月份全国报表
from openpyxl.styles import colors, Font, Fill, NamedStyle
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl import load_workbook
import pandas as pd
import numpy as np
import pymongo
import pymysql
import datetime
import time
import os
from copy import copy
import requests
from urllib3 import encode_multipart_formdata
from functools import reduce
from openpyxl.formatting.rule import DataBarRule
import warnings
warnings.filterwarnings("ignore")

print('开始时间{}'.format(time.strftime("%Y-%m-%d-%H:%M:%S", time.localtime())))
# 连接mongdb
conn = pymongo.MongoClient('mongodb://{}:{}@{}:{}/?authSource={}'
                           .format("cda_epmall", "3PUkMsWCTu424Renrffb",
                                   "dds-wz99d33fe2f72610a042-pub.mongodb.rds.aliyuncs.com",
                                   "3717", "epmall"))
conn_epmall = conn["epmall"]


df_order_all = conn_epmall['epmall_orders_100063']  # 全量订单表

df_order_all = pd.DataFrame(list(df_order_all.find({'state': {'$gte': 1, '$lte': 9}},
                                                   {'_id': 0, 'state': 1, 'order_no': 1, 'create_date': 1, 'user_id': 1,
                                                    'nick': 1, 'mobile': 1,
                                                    'isback': 1, 'commodity_id': 1, 'commodity_name': 1,
                                                    'order_type': 1,
                                                    'recommend_store_id': 1, 'third_platform': 1, 'is_selfpick': 1,
                                                    'pick_store_id': 1, 'amount_total': 1, 'totalDiscount': 1,
                                                    'recharge_balance': 1})))
df_order_all['recommend_store_id'] = df_order_all['recommend_store_id'].mask(df_order_all['order_type'] == 0,
                                                                             df_order_all['pick_store_id'])
df_order_all.replace('', '平台', inplace=True)
df_order_all['recommend_store_id'].fillna('平台', inplace=True)

df_order_all0 = conn_epmall['epmall_orders_100063']  # 全量订单表 退款找订单id用

df_order_all0 = pd.DataFrame(list(df_order_all0.find({'state': {'$gte': 1, '$lte': 9}},
                                                     {'_id': 0, 'state': 1, 'order_no': 1, 'create_date': 1,
                                                      'user_id': 1, 'nick': 1, 'mobile': 1,
                                                      'isback': 1, 'commodity_id': 1, 'commodity_name': 1,
                                                      'order_type': 1,
                                                      'recommend_store_id': 1, 'third_platform': 1, 'is_selfpick': 1,
                                                      'pick_store_id': 1, 'amount_total': 1, 'totalDiscount': 1,
                                                      'recharge_balance': 1})))
df_order_all0['recommend_store_id'] = df_order_all0['recommend_store_id'].mask(df_order_all0['order_type'] == 0,
                                                                               df_order_all0['pick_store_id'])
df_order_all0.replace('', '平台', inplace=True)
df_order_all0['recommend_store_id'].fillna('平台', inplace=True)

df_recharge_all = conn_epmall['epmall_recharge_log_100063']  # 全量充值表
df_recharge_all = pd.DataFrame(
    list(df_recharge_all.find({'pay_state': 1}, {'_id': 0, 'activity_store_id': 1, 'deal_no': 1})))
df_recharge_all.replace('', '平台', inplace=True)
df_recharge_all['activity_store_id'].fillna('平台', inplace=True)

x = datetime.date.today()
oneday = datetime.timedelta(days=1)
y = x - oneday
str_month = datetime.datetime(y.year, y.month, 1)

str_date = str(str_month)[:10]

en_date = str(y)
date=str(x)

begin_date = str(time.mktime(time.strptime(str_date, "%Y-%m-%d")))

end_date = str(time.mktime(time.strptime(date, "%Y-%m-%d")))


date_list = []
while True:
    date_list.append(str(x))
    if str(x) == str(str_month)[:10]:
        break
    x = x - oneday

print('开始时间:', begin_date)
print('结束时间:', end_date)

col_order = conn_epmall['epmall_orders_100063']  # 订单表

df_order = pd.DataFrame(
    list(col_order.find({'create_date': {'$gt': begin_date, '$lt': end_date}, 'state': {'$gte': 1, '$lte': 9}},
                        {'_id': 0, 'state': 1, 'order_no': 1, 'create_date': 1, 'user_id': 1, 'nick': 1, 'mobile': 1,
                         'isback': 1, 'commodity_id': 1, 'commodity_name': 1, 'order_type': 1,
                         'recommend_store_id': 1, 'third_platform': 1, 'is_selfpick': 1, 'pick_store_id': 1,
                         'amount_total': 1, 'totalDiscount': 1, 'recharge_balance': 1})))
df_order.eval(""" accounted_money= amount_total - totalDiscount - recharge_balance""", inplace=True)
df_order['accounted_money'] = df_order['accounted_money'] / 100
df_order['recommend_store_id'] = df_order['recommend_store_id'].mask(df_order['order_type'] == 0,
                                                                     df_order['pick_store_id'])
df_order.replace('', '平台', inplace=True)
df_order['recommend_store_id'].fillna('平台', inplace=True)
df_order = df_order[df_order['accounted_money'] > 0]


col_app_user = conn_epmall['epmall_app_user_100063']  # 小程序用户表
df_app_user = pd.DataFrame(list(col_app_user.find({},
                                                  {'_id': 0, 'user_id': 1, 'promote_orders': 1,
                                                   'promote_orders_time': 1, 'repurchase_orders': 1,
                                                   'repurchase_orders_time': 1, 'channel_belong_name': 1,
                                                   'deliver_id': 1})))  # 用户id,是否升单,升单时间,是否复购,复购时间


col_serve = conn_epmall['epmall_service_record_100063']  # 服务记录表
df_serve = pd.DataFrame(list(col_serve.find({'create_time': {'$gt': begin_date, '$lt': end_date}},
                                            {'_id': 0, 'user_id': 1, 'store_id': 1, 'creat_time': 1,
                                             'afterDiscount': 1})))
df_serve['store_id'].fillna('平台', inplace=True)
df_serve['afterDiscount'] = df_serve['afterDiscount'] / 100


col_recharge = conn_epmall['epmall_recharge_log_100063']  # 充值表
df_recharge = pd.DataFrame(list(
    col_recharge.find({'recharge_date': {'$gt': begin_date, '$lt': end_date}, 'pay_state': 1},
                      {'_id': 0, 'user_id': 1, 'amount_pay': 1, 'activity_store_id': 1, 'deal_no': 1,
                       'recharge_date': 1})))
df_recharge['amount_pay'] = df_recharge['amount_pay'] / 100
df_recharge['activity_store_id'].fillna('平台', inplace=True)

df_recharge['user_id'] = pd.to_numeric(df_recharge['user_id'], errors='raise')


col_refund = conn_epmall['epmall_refund_100063']  # 充值退款表  'create_time':{'$gt':begin_date,'$lt':end_date},
df_refund = pd.DataFrame(list(col_refund.find({'refund_time': {'$gt': begin_date, '$lt': end_date}, 'state': 2},
                                              {'_id': 0, 'user_id': 1, 'refund_money': 1, 'deal_no': 1})))
if len(df_refund) != 0:
    df_refund['refund_money'] = df_refund['refund_money'] / 100
    df_refund = df_refund.rename(columns={'refund_money': 'amount_refund'})




col_back=conn_epmall['epmall_back_order']  #退单表
df_back = pd.DataFrame(list(col_back.find({'cashier_data':{'$gt':begin_date,'$lt':end_date},'back_result':4},{'_id':0,'order_no':1,'user_id':1,'recharge_balance':1,'commodity_id':1,'amount_refund':1})))
if  len(df_back)!=0:
    df_back['wxrefund_money']=df_back['amount_refund'] - df_back['recharge_balance']
    df_back['wxrefund_money']=df_back['wxrefund_money'] / 100
    df_back=df_back[df_back['wxrefund_money']>0]


# ### 卖卡数量 卖卡金额

col_commodity = conn_epmall['epmall_commodity_100063']  # 商品表
df_commodity = pd.DataFrame(list(col_commodity.find({},
                                                    {'_id': 0, 'commodity_id': 1,
                                                     'variety_id': 1})))

variety_id = ['1646038407932', '1646209380573', '1646491684554','1650421003620']
df_commodity = df_commodity[df_commodity['variety_id'].isin(variety_id)]
commodity_id_list = list(df_commodity['commodity_id'])  # 卡商品id

df_order_all = df_order_all[df_order_all['commodity_id'].isin(commodity_id_list)]  # 此为全量售卡订单

order_card = df_order[df_order['commodity_id'].isin(commodity_id_list)]
# 此为售卡订单
len(order_card)

# 算人数 来自app订单和来自微信小程序要去重
third_platform1 = [0, 3, 5, 6]
order_card1 = order_card[order_card['third_platform'].isin(third_platform1)].drop_duplicates(subset=['user_id'])
len(order_card1)

# 拼接
third_platform2 = [1, 2, 4]
order_card2 = order_card[order_card['third_platform'].isin(third_platform2)]
order_card_new = pd.concat([order_card1, order_card2])


## 卖卡人数 卖卡金额

card_sum = order_card_new.groupby('recommend_store_id')['user_id'].count().reset_index()

card_sum = card_sum.rename(columns={'user_id': '售卡人数'})
card_sum.sum()

card_money = order_card_new.groupby('recommend_store_id')['accounted_money'].sum().reset_index()

card_money = card_money.rename(columns={'accounted_money': '售卡金额'})

# 计算升单金额及人数、卡升单金额及人数、复购金额及人数时，需增加体验卡剔除的机制。
df_order_remove = df_order[~df_order['commodity_id'].isin(commodity_id_list)]

# ### 升单人数 ,升单金额

# 升单人数 (有app_user 表里 promote_orders =1 都算) 当月
# 升单金额(有app_user 表里 promote_orders =1 ,升单订单的筛选方法:promote_orders_time<=订单时间<repurchase_orders_time)

col_user_up = conn_epmall['epmall_app_user_100063']  # 小程序用户表promote_orders_time 为当月的
df_user_up = pd.DataFrame(list(col_user_up.find({'promote_orders_time': {'$gt': begin_date, '$lt': end_date}},
                                                {'_id': 0, 'user_id': 1, 'promote_orders': 1, 'promote_orders_time': 1,
                                                 'repurchase_orders': 1,
                                                 'repurchase_orders_time': 1})))

df_promote_orders = df_user_up[df_user_up['promote_orders'] == 1]

order_up = pd.merge(df_promote_orders,df_order,on='user_id')
recharge_up = pd.merge(df_promote_orders,df_recharge,on='user_id')

recharge_up=recharge_up.loc[:,['activity_store_id','user_id']]
recharge_up.rename(columns={'activity_store_id':'recommend_store_id'},inplace=True)

up_sum=order_up[order_up['accounted_money']!=0]
up_sum=up_sum.loc[:,['recommend_store_id','user_id']]
up_sum=pd.concat([up_sum,recharge_up])
up_sum=up_sum.drop_duplicates(subset='user_id')

up_sum=up_sum.groupby('recommend_store_id')['user_id'].count().reset_index()#升单人数
up_sum=up_sum.rename(columns={'user_id':'升单人数'})

df_user_up1 = pd.merge(df_user_up, df_order_remove, on='user_id', how='inner')  # 合并订单和user
df_user_up2 = pd.merge(df_user_up, df_recharge, on='user_id', how='inner')  # 合并充值和user

# 计算升单订单金额
df_user_promote = df_user_up1[df_user_up1['repurchase_orders'] == 1]
df_user_promote = df_user_promote[(df_user_promote['create_date'] < df_user_promote['repurchase_orders_time']) & (
            df_user_promote['promote_orders_time'] <= df_user_promote['create_date'])]

df_user_promote0 = df_user_up1[df_user_up1['repurchase_orders'] == 0]
df_user_promote0 = df_user_promote0[(df_user_promote0['promote_orders_time'] <= df_user_promote0['create_date'])]

df_user_promote0 = pd.concat([df_user_promote0, df_user_promote])

df_user_promote0 = df_user_promote0.loc[:, ['user_id', 'accounted_money', 'recommend_store_id']]

# 计算充值金额
df_user_repurchase = df_user_up2[df_user_up2['repurchase_orders'] == 1]
df_user_repurchase = df_user_repurchase[
    (df_user_repurchase['recharge_date'] < df_user_repurchase['repurchase_orders_time']) & (
                df_user_repurchase['promote_orders_time'] <= df_user_repurchase['recharge_date'])]

df_user_repurchase0 = df_user_up2[df_user_up2['repurchase_orders'] == 0]
df_user_repurchase0 = df_user_repurchase0[
    (df_user_repurchase0['promote_orders_time'] <= df_user_repurchase0['recharge_date'])]
df_user_repurchase0 = pd.concat([df_user_repurchase0, df_user_repurchase])

df_user_repurchase0.rename(columns={'activity_store_id': 'recommend_store_id', 'amount_pay': 'accounted_money'},
                           inplace=True)

df_user_repurchase0 = df_user_repurchase0.loc[:, ['user_id', 'accounted_money', 'recommend_store_id']]

# In[305]:


up_money = pd.concat([df_user_repurchase0, df_user_promote0])

# In[306]:


up_money = up_money.groupby('recommend_store_id')['accounted_money'].sum().reset_index()
up_money = up_money.rename(columns={'accounted_money': '升单金额'})  # 计算升单金额

# ### 卡到店人数,卡升单人数,卡升单金额,卡升单转化率 ,卡升单客单价

# 卡到店人数 :先筛选出本月购卡者 然后与服务记录表相关联订单
#
# 卡升单人数 先筛选出本月购卡者 查找 app_user 表里 promote_orders =1 且promote_orders_time 在购卡订单时间之后者
#
# 卡升单金额 卡升单客户的升单订单实付金额 (升单订单的筛选方法:promote_orders_time<=订单时间<repurchase_orders_time)
# 卡升单转化率 =卡升单人数/卡到店人数
#
# 卡升单客单价 =卡升单金额/卡升单人数

# 卡到店人数
card_shop = pd.merge(df_order_all, df_serve, on='user_id', how='left')
card_shop = card_shop.drop_duplicates('user_id')

card_shop_sum = card_shop.groupby('store_id')['user_id'].count().reset_index()

card_shop_sum = card_shop_sum.rename(columns={'user_id': '卡到店人数'})
card_shop_sum.rename(columns={'store_id': 'recommend_store_id'}, inplace=True)

card_user_up = conn_epmall['epmall_app_user_100063']  # 小程序用户表promote_orders_time 为当月的
card_user_up = pd.DataFrame(list(card_user_up.find({'promote_orders_time': {'$gt': begin_date, '$lt': end_date}},
                                                   {'_id': 0, 'user_id': 1, 'promote_orders': 1,
                                                    'promote_orders_time': 1, 'repurchase_orders': 1,
                                                    'repurchase_orders_time': 1})))

# 卡升单人数  当月购卡并且当月升单


# ### 卡升单金额

card_up_sum2 = pd.merge(df_order_all['user_id'], card_user_up, on='user_id', how='inner')
card_up_sum2 = pd.merge(card_up_sum2, df_order, on='user_id', how='inner')
card_up_sum2 = card_up_sum2[~card_up_sum2['order_no'].isin(list(order_card['order_no']))]
card_up_sum2 = card_up_sum2.drop_duplicates(subset='order_no')


card_up_sum3 = pd.merge(df_order_all['user_id'], card_user_up, on='user_id', how='inner')
card_up_sum3 = pd.merge(card_up_sum3, df_recharge, on='user_id', how='inner')
card_up_sum3 = card_up_sum3.drop_duplicates(subset='deal_no')

# #计算卡升单订单金额
df_up_orders = card_up_sum2[card_up_sum2['repurchase_orders'] == 1]
df_up_orders = df_up_orders[(df_up_orders['create_date'] < df_up_orders['repurchase_orders_time']) & (
            df_up_orders['promote_orders_time'] <= df_up_orders['create_date'])]

df_up_orders0 = card_up_sum2[card_up_sum2['repurchase_orders'] == 0]
df_up_orders0 = df_up_orders0[(df_up_orders0['promote_orders_time'] <= df_up_orders0['create_date'])]

df_up_orders0 = pd.concat([df_up_orders0, df_up_orders0])

df_up_orders0 = df_up_orders0.drop_duplicates(subset='order_no')

df_up_orders0 = df_up_orders0.loc[:, ['user_id', 'accounted_money', 'recommend_store_id']]

# 计算卡升单充值金额
df_up_recharge = card_up_sum3[card_up_sum3['repurchase_orders'] == 1]
df_up_recharge = df_up_recharge[(df_up_recharge['recharge_date'] <
                                 df_up_recharge['repurchase_orders_time']) & (df_up_recharge['promote_orders_time']
                                                                              <= df_up_recharge['recharge_date'])]

df_up_recharge0 = card_up_sum3[card_up_sum3['repurchase_orders'] == 0]
df_up_recharge0 = df_up_recharge0[(df_up_recharge0['promote_orders_time'] <= df_up_recharge0['recharge_date'])]
df_up_recharge0 = pd.concat([df_up_recharge0, df_up_recharge])

df_up_recharge0.rename(columns={'activity_store_id': 'recommend_store_id', 'amount_pay': 'accounted_money'},
                       inplace=True)

df_up_recharge0 = df_up_recharge0.loc[:, ['user_id', 'accounted_money', 'recommend_store_id']]

card_up_money = pd.concat([df_up_recharge0, df_up_orders0])




card_up_money1 = card_up_money.groupby('recommend_store_id')['accounted_money'].sum().reset_index()
card_up_money1 = card_up_money1.rename(columns={'accounted_money': '卡升单金额'})

# 卡升单金额


# ### 卡升单人数

card_up_sum = card_up_money.drop_duplicates('user_id')

card_up_sum = card_up_sum.groupby('recommend_store_id')['user_id'].count().reset_index()
card_up_sum = card_up_sum.rename(columns={'user_id': '卡升单人数'})
card_up_sum.head(50)

# ### 服务人数,服务金额

# 服务人数
serve_sum = df_serve.drop_duplicates('user_id')
serve_sum = serve_sum.groupby('store_id')['user_id'].count().reset_index()
serve_sum.rename(columns={'user_id': '服务人数', 'store_id': 'recommend_store_id'}, inplace=True)

# 服务金额
serve_money = df_serve.groupby('store_id')['afterDiscount'].sum().reset_index()
serve_money.rename(columns={'afterDiscount': '服务金额', 'store_id': 'recommend_store_id'}, inplace=True)


# ### 老客到店人数,复购成交人数,复购金额

def f(x):
    return datetime.datetime.fromtimestamp(int(x)).replace(hour=0, minute=0, second=0, microsecond=0)

# 老客到店人数
col_serve1 = conn_epmall['epmall_service_record_100063']  # 服务记录表
df_serve1 = pd.DataFrame(list(col_serve1.find({'create_time': {'$gt': begin_date, '$lt': end_date}},
                                              {'_id': 0, 'user_id': 1, 'store_id': 1, 'create_time': 1})))
df_serve1['store_id'].fillna('平台', inplace=True)
repurchase1 = df_app_user[df_app_user['promote_orders'] == 1]
repurchase_sum = pd.merge(df_serve1, repurchase1, on='user_id', how='inner')
repurchase_sum = repurchase_sum[repurchase_sum['promote_orders_time'] < repurchase_sum['create_time']]
repurchase_sum['create_time'] = repurchase_sum['create_time'].apply(lambda x: x[0:10])
repurchase_sum['promote_orders_time'] = repurchase_sum['promote_orders_time'].apply(lambda x: x[0:10])
repurchase_sum = repurchase_sum[repurchase_sum['create_time'].apply(f) != repurchase_sum['promote_orders_time'].apply(f)]

repurchase_sum1 = repurchase_sum.drop_duplicates('user_id')

repurchase_sum1 = repurchase_sum1.groupby('store_id')['user_id'].count().reset_index()
repurchase_sum1 = repurchase_sum1.rename(columns={'user_id': '老客到店人数', 'store_id': 'recommend_store_id'})


df_serve1['create_time'] = df_serve1['create_time'].apply(lambda x: x[0:10])

# ### 新客到店人数

repurchase_sum2 = df_serve1[~df_serve1['create_time'].isin(list(repurchase_sum['create_time']))]

repurchase_sum2.drop_duplicates('user_id', inplace=True)

repurchase_sum2 = repurchase_sum2.groupby('store_id')['user_id'].count().reset_index()
repurchase_sum2 = repurchase_sum2.rename(columns={'user_id': '新客到店人数', 'store_id': 'recommend_store_id'})
repurchase_sum2.sum()

# ### 复购成交人数

repurchase = df_app_user[df_app_user['repurchase_orders'] == 1]
repurchase_people = pd.merge(df_order_remove, repurchase, on='user_id', how='inner')
repurchase_people = repurchase_people.drop_duplicates('user_id')


repurchase_people = repurchase_people.groupby('recommend_store_id')['user_id'].count().reset_index()
repurchase_people = repurchase_people.rename(columns={'user_id': '复购成交人数'})

df_complex1 = pd.merge(df_app_user, df_order_remove, on='user_id', how='inner')  # 合并订单和用户表
df_complex2 = pd.merge(df_app_user, df_recharge, on='user_id', how='inner')  # 合并充值和用户表
df_complex1 = df_complex1[df_complex1['repurchase_orders'] == 1]
df_complex2 = df_complex2[df_complex2['repurchase_orders'] == 1]

# 计算复购订单金额

df_complex1 = df_complex1[df_complex1['repurchase_orders_time'] <= df_complex1['create_date']]

df_complex2 = df_complex2[df_complex2['repurchase_orders_time'] <= df_complex2['recharge_date']]
df_complex2.rename(columns={'amount_pay': 'accounted_money', 'activity_store_id': 'recommend_store_id'}, inplace=True)


df_complex1 = df_complex1.loc[:, ['user_id', 'accounted_money', 'recommend_store_id']]
df_complex2 = df_complex2.loc[:, ['user_id', 'accounted_money', 'recommend_store_id']]

repurchase_money = pd.concat([df_complex2, df_complex1])

# 复购成交金额

repurchase_money = repurchase_money.groupby('recommend_store_id')['accounted_money'].sum().reset_index()
repurchase_money = repurchase_money.rename(columns={'accounted_money': '复购成交金额'})

# ### 商品购买人数,商品订单金额

commodity_order = df_order[df_order['order_type'] == 0]

commodity_order_sum = commodity_order.groupby('recommend_store_id')['accounted_money'].count().reset_index()
commodity_order_sum.rename(columns={'accounted_money': '商品购买人数'}, inplace=True)

commodity_order_money = commodity_order.groupby('recommend_store_id')['accounted_money'].sum().reset_index()
commodity_order_money.rename(columns={'accounted_money': '商品订单金额'}, inplace=True)

# ### 服务购买人数,服务订单金额

service_order = df_order[df_order['order_type'] != 0]

service_order_sum = service_order.groupby('recommend_store_id')['accounted_money'].count().reset_index()
service_order_sum.rename(columns={'accounted_money': '服务购买人数'}, inplace=True)

service_order_money = service_order.groupby('recommend_store_id')['accounted_money'].sum().reset_index()
service_order_money.rename(columns={'accounted_money': '服务订单金额'}, inplace=True)

# ### 充值人数,充值金额,总成交金额,总退款金额

###总充值金额 充值人数

recharge_sum = df_recharge.drop_duplicates('user_id')
recharge_sum = df_recharge.groupby('activity_store_id')['user_id'].count().reset_index()
recharge_sum = recharge_sum.rename(columns={'activity_store_id': 'recommend_store_id', 'user_id': '充值人数'})

#recharge_sum.to_excel('充值金额.xlsx')
recharge_pay = df_recharge.groupby('activity_store_id')['amount_pay'].sum().reset_index()
recharge_pay = recharge_pay.rename(columns={'activity_store_id': 'recommend_store_id', 'amount_pay': '充值金额'})

# 总订单金额

order_pay = df_order.groupby('recommend_store_id')['accounted_money'].sum().reset_index()
order_pay = order_pay.rename(columns={'accounted_money': '总订单金额'})

##### 总成交金额
total_pay = pd.merge(order_pay, recharge_pay, on='recommend_store_id', how='outer')
total_pay.fillna(0, inplace=True)
total_pay.eval(""" 总成交金额= 总订单金额 + 充值金额 """, inplace=True)
total_pay = total_pay.loc[:, ['recommend_store_id', '总成交金额']]

# 总成交人数

total_sum = df_order.drop_duplicates('user_id')

total_sum = total_sum.groupby('recommend_store_id')['user_id'].count().reset_index()
total_sum = total_sum.rename(columns={'user_id': '总成交人数'})

# ### 自拓客

plan = conn_epmall['epmall_deliver_plan_100063']  # 客户渠道表
plan = pd.DataFrame(list(plan.find({}, {'_id': 0, 'deliver_page.link_id': 1, 'deliver_id': 1})))
plan['deliver_page'] = plan['deliver_page'].apply(lambda x: str(x)[13:-2])

# 自拓客到店人数
plan_sum = pd.merge(df_serve, df_app_user, how='left')
plan_sum = pd.merge(plan_sum, plan, on='deliver_id', how='left')
plan_sum = plan_sum[plan_sum['store_id'] == plan_sum['deliver_page']]
plan_sum.drop_duplicates('user_id', inplace=True)
plan_sum = plan_sum.groupby('store_id')['user_id'].count().reset_index()
plan_sum = plan_sum.rename(columns={'user_id': '自拓客到店人数', 'store_id': 'recommend_store_id'})

# 自拓客实收金额
#自拓客实收订单金额
plan_money=pd.merge(df_order,df_app_user,how='left')
plan_money=pd.merge(plan_money,plan,on='deliver_id',how='left')
plan_money=plan_money[plan_money['recommend_store_id']==plan_money['deliver_page']]

#自拓客实收充值金额
plan_money1=pd.merge(df_recharge,df_app_user,how='left')
plan_money1=pd.merge(plan_money1,plan,on='deliver_id',how='left')
plan_money1=plan_money1[plan_money1['activity_store_id']==plan_money1['deliver_page']]

plan_money1.rename(columns={'amount_pay':'accounted_money','activity_store_id':'recommend_store_id'},inplace=True)
plan_money=plan_money.loc[:,['user_id','accounted_money','recommend_store_id']]
plan_money1=plan_money1.loc[:,['user_id','accounted_money','recommend_store_id']]

#自拓客实收金额=自拓客实收充值金额+自拓客实收订单金额
plan_money2=pd.concat([plan_money,plan_money1])
plan_money2=plan_money2.groupby('recommend_store_id')['accounted_money'].sum().reset_index()
plan_money2.rename(columns={'accounted_money':'自拓客实收金额'},inplace=True)

# 自拓客成交人数
plan_money_sum = pd.merge(df_order, df_app_user, how='left')
plan_money_sum = pd.merge(plan_money_sum, plan, on='deliver_id', how='left')
plan_money_sum = plan_money_sum[plan_money_sum['recommend_store_id'] == plan_money_sum['deliver_page']]
plan_money_sum.drop_duplicates('user_id', inplace=True)
plan_money_sum = plan_money_sum.groupby('recommend_store_id')['user_id'].count().reset_index()
plan_money_sum.rename(columns={'user_id': '自拓客成交人数'}, inplace=True)

# ### 总退款金额
df_back_new = None
df_refund_new = None

# #充值退款
if len(df_refund) != 0:
    df_refund_new = pd.merge(df_refund, df_recharge_all, on='deal_no', how='left')

    df_refund_new = df_refund_new.groupby('activity_store_id')['amount_refund'].sum().reset_index()
    df_refund_new = df_refund_new.rename(columns={'activity_store_id': 'recommend_store_id'})

if len(df_back) != 0:
    df_back_new = pd.merge(df_back, df_order_all0, on='order_no', how='left')  # 订单退款

    df_back_new = df_back_new.loc[:, ['order_no', 'recommend_store_id', 'wxrefund_money']]
    df_back_new = df_back_new.groupby('recommend_store_id')['wxrefund_money'].sum().reset_index()

total_money = None
if len(df_back) != 0 and len(df_refund) != 0:
    total_money = pd.merge(df_back_new, df_refund_new, on='recommend_store_id', how='outer')
    total_money.fillna(0, inplace=True)
    total_money.eval(""" 总退款金额= wxrefund_money + amount_refund """, inplace=True)
    total_money = total_money.loc[:, ['recommend_store_id', '总退款金额']]

if len(df_back) == 0 and len(df_refund) != 0:
    total_money = df_refund_new.rename(columns={'amount_refund': '总退款金额'})
if len(df_refund) == 0 and len(df_back) != 0:
    total_money = df_back_new.rename(columns={'wxrefund_money': '总退款金额'})

# 体验卡退款commodity_id_list


df_back_yes=None
if len(df_back)!=0:
    df_back_yes=df_back[df_back['commodity_id'].isin(commodity_id_list)]#筛选体验卡退款
    df_back_yes=pd.merge(df_back_yes,df_order_all0,on='order_no',how='left')#订单退款

    df_back_yes=df_back_yes.loc[:,['order_no','recommend_store_id','wxrefund_money']]
    df_back_yes=df_back_yes.groupby('recommend_store_id').agg({'order_no':'count','wxrefund_money':'sum'}).reset_index()
    df_back_yes = df_back_yes.rename(columns={'wxrefund_money': '体验卡退款金额','order_no': '体验卡退款数量'})



# ### 合并

if len(df_refund) != 0 or len(df_back) != 0:
    dfs = [total_sum, total_pay, total_money, recharge_sum, recharge_pay, repurchase_money, repurchase_people,
           repurchase_sum1, repurchase_sum2, card_money, card_sum, up_money, up_sum, commodity_order_sum,
           commodity_order_money, service_order_sum, service_order_money, serve_sum, serve_money, plan_sum,
           plan_money_sum, plan_money2]
else:
    dfs = [total_sum, total_pay, recharge_sum, recharge_pay, repurchase_money, repurchase_people, repurchase_sum1,
           repurchase_sum2, card_money, card_sum, up_money, up_sum, commodity_order_sum, commodity_order_money,
           service_order_sum, service_order_money, serve_sum, serve_money, plan_sum, plan_money_sum, plan_money2]

if df_back_yes is not None:
    dfs.append(df_back_yes)



df_result = reduce(lambda left, right: pd.merge(left, right, on='recommend_store_id', how='outer'), dfs)
platform = df_result[df_result['recommend_store_id'] == '平台']
platform['store_name'] = '平台'




df_store = conn_epmall['epmall_store_100063']  # 门店

df_store = pd.DataFrame(list(df_store.find({}, {'_id': 0, 'store_id': 1,'store_name': 1,'city':1})))
df_result = pd.merge(df_result, df_store, left_on='recommend_store_id', right_on='store_id', how='inner')
df_result.drop('store_id', axis=1, inplace=True)
df_result = pd.concat([platform, df_result])
df_result = df_result.rename(columns={'city': '城市'})

df_result.fillna(0, inplace=True)

if len(df_refund) != 0 or len(df_back) != 0:
    df_result.eval("""
    退款率 = 总退款金额 / 总成交金额
    售卡金额占比 = 售卡金额 / 总成交金额
    复购率 = 复购成交人数/老客到店人数
    复购金额占比 = 复购成交金额/总成交金额
    实际成交金额 = 总成交金额-总退款金额
    充值占比 = 充值金额/总成交金额
    商品订单金额占比 = 商品订单金额/总成交金额
    服务订单金额占比 = 服务订单金额/总成交金额
    自拓客金额占比=自拓客实收金额/总成交金额
    升单率 = 升单人数/新客到店人数
    复购客单价=复购成交金额/复购成交人数
    升单客单价=升单金额/升单人数
    升单金额占比 = 升单金额/总成交金额""", inplace=True)
else:
    df_result.eval("""
        退款率 = 0
        总退款金额=0
        售卡金额占比 = 售卡金额 / 总成交金额
        复购率 = 复购成交人数/老客到店人数
        复购金额占比 = 复购成交金额/总成交金额
        实际成交金额 = 总成交金额-总退款金额
        充值占比 = 充值金额/总成交金额
        商品订单金额占比 = 商品订单金额/总成交金额
        服务订单金额占比 = 服务订单金额/总成交金额
        自拓客金额占比=自拓客实收金额/总成交金额
        升单率 = 升单人数/新客到店人数
        复购客单价=复购成交金额/复购成交人数
        升单客单价=升单金额/升单人数
        升单金额占比 = 升单金额/总成交金额 """, inplace=True)


df_result['store_name'] = df_result['store_name'].apply(lambda x: x.strip())

##链接目标业绩
db = pymysql.connect(host='l5vhlstu.ipyingshe.net', database='sys', user='root', password='root123', port=31453,charset='utf8')
sql = 'select * from July'
df_region = pd.read_sql(sql, db)

df_result = pd.merge(df_region, df_result, left_on='门店ID', right_on='recommend_store_id', how='right')

df_result['责任目标']=df_result['责任目标'].apply(lambda x: 0 if x is np.nan else int(x))
df_result['冲刺目标']=df_result['冲刺目标'].apply(lambda x: 0 if x is np.nan else int(x))
df_result['超越目标']=df_result['超越目标'].apply(lambda x: 0 if x is np.nan else int(x))


# df_result=df_result.astype({'责任目标':'int','冲刺目标':'int','超越目标':'int'})
if len(df_back)==0:
    df_result['体验卡退款金额']=0
    df_result['体验卡退款数量'] = 0
df_result['目标完成进度（责任）']=df_result['总成交金额']/df_result['责任目标']
df_result['目标完成进度（冲刺）']=df_result['总成交金额']/df_result['冲刺目标']
df_result['目标完成进度（超越）']=df_result['总成交金额']/df_result['超越目标']
df_result = df_result.loc[:,
            ['大区', '门店名称','责任目标','冲刺目标','超越目标', '城市','总成交人数', '总成交金额', '总退款金额','体验卡退款金额','体验卡退款数量', '实际成交金额', '退款率', '充值人数', '充值金额', '充值占比', '商品购买人数',
             '商品订单金额', '商品订单金额占比',
             '服务购买人数', '服务订单金额', '服务订单金额占比', '售卡人数', '售卡金额', '售卡金额占比', '新客到店人数', '升单人数', '升单金额','升单客单价', '升单金额占比', '升单率',
             '老客到店人数', '复购成交人数', '复购成交金额','复购客单价', '复购金额占比', '复购率', '服务人数', '服务金额', '自拓客到店人数', '自拓客成交人数',
             '自拓客实收金额', '自拓客金额占比','目标完成进度（责任）','目标完成进度（冲刺）','目标完成进度（超越）']]



df_result.sort_values("大区", inplace=True)
df_result=df_result.fillna(0)
df_result.replace(np.inf, 0, inplace=True)

df_result.to_excel( '全国市场门店销售月报表临时用.xlsx', index=False, header=True)

area_list1=['平台', '广东江苏河南区域', '浙江区域', '西南区域', '武汉-北京']
concat_data=pd.DataFrame({})
for area in area_list1:
    temp_df=df_result.loc[df_result['大区']==area,:]
    temp_df.insert(0, '序号', range(1,len(temp_df)+1))
    concat_data=pd.concat((temp_df,concat_data))


concat_data1=concat_data.loc[:,['序号','大区','门店名称','责任目标','冲刺目标','超越目标','售卡人数','售卡金额','售卡金额占比','新客到店人数','升单人数','升单金额','升单客单价','升单率','升单金额占比','老客到店人数', '复购成交人数',
      '复购成交金额','复购客单价','复购率','复购金额占比','自拓客到店人数', '自拓客成交人数','自拓客实收金额', '自拓客金额占比','商品购买人数',
      '商品订单金额', '商品订单金额占比','总成交人数', '总成交金额', '总退款金额','实际成交金额','服务人数', '服务金额','退款率' ,'目标完成进度（责任）','目标完成进度（冲刺）','目标完成进度（超越）']]


# 将时间戳转化为时间格式
def time_t(x):
    if x=='':
        return ''
    elif x==' ':
        return ''
    elif str(x)=='nan':
        return ''
    elif str(x)=='NaN':
        return ''
    else:
        x=int(x)
        time_local = time.localtime(x/1000)
        create_date = time.strftime('%Y-%m-%d %H:%M:%S', time_local)
        return create_date

def getYesterday():
    today = datetime.date.today()
    oneday = datetime.timedelta(days=1)
    yesterday = today - oneday
    return yesterday

date =str(getYesterday())
month1=date[:7]


concat_data1.to_excel( './全国市场门店销售报表月度累计'+date+'.xlsx', index=False)


df_area=pd.read_excel('区域城市对照表.xlsx')

df_city_info=pd.read_excel(r'D:\python_project\projects\region_adverting\city_consume.xlsx')
df_city_info=df_city_info.loc[:,['城市','广告总额','进线人数']]
df_city_info.insert(0,'日期',date)

df_exist=pd.read_excel(r'各城市广告费累计.xlsx')

df_concat_advinfo=pd.concat([df_exist,df_city_info])
df_concat_advinfo.astype({'日期':'str','城市':'str','广告总额':'float','进线人数':'int'})
df_concat_advinfo.drop_duplicates(subset=['日期','城市'],keep='last',inplace=True)
total_city_adv=df_concat_advinfo.loc[df_concat_advinfo['日期'].str.contains(month1),['城市','广告总额']]
# total_city_adv=df_concat_advinfo.loc[:,['城市','广告总额']]

total_adv=total_city_adv.groupby('城市').sum().reset_index()
total_adv_amount=total_adv['广告总额'].sum()

total_area_adv=pd.merge(total_adv,df_area,on='城市',how='left')
total_area_adv=total_area_adv.loc[:,['大区','广告总额']]

total_area_adv=total_area_adv.groupby('大区').sum()

adv_amount_dict=total_area_adv.to_dict()
area_amount=adv_amount_dict['广告总额']
df_concat_advinfo.to_excel('各城市广告费累计.xlsx',index=False)

wb = load_workbook(filename='./全国市场门店销售报表月度累计'+date+'.xlsx')
ws = wb.active


for col in  ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in col:
            cell.alignment = Alignment(horizontal='center', vertical='center')


prj = ws.columns
place_list=[]
prjTuple = tuple(prj)
for idx in range(1,2):# 遍历A、B、C、D列  #(0,len(prjTuple)) 遍历全部列
    for cell in prjTuple[idx]:
        if (cell.value !='大区' and cell.value is not None):

            place_list.append(cell.value)


place_dict= {}
for i in place_list:
    place_dict[i] = place_list.count(i)


place_list1=[]
for j in place_list:
    if j not in place_list1:
        place_list1.append(j)

font1 = Font(name='宋体', size=11, b=True)
font2 = Font(name='宋体', size=14, b=True)
font3 = Font(name='宋体', size=24, b=True)

#边框
line_t = Side(style='thin', color='000000')  # 细边框
line_m = Side(style='medium', color='000000')  # 粗边框
border1 = Border(top=line_t, bottom=line_t, left=line_t, right=line_t)
# 与标题相邻的边设置与标题一样
border2 = Border(top=line_t, bottom=line_t, left=line_t, right=line_t)

# 填充
fill = PatternFill('solid', fgColor='CFCFCF')#无色
fill1=PatternFill('solid', fgColor='FFCC33')#浅黄色
fill2=PatternFill('solid', fgColor='FF0000')#红色
fill3=PatternFill('solid', fgColor='669966')#绿色
fill4=PatternFill('solid', fgColor='D2691E')#橙色
fill5=PatternFill('solid', fgColor='336699')#蓝色
fill6=PatternFill('solid', fgColor='CC9999')#粉色
fill7=PatternFill('solid', fgColor='CC9966')#褐色
fill10=PatternFill('solid', fgColor='FFFF00')#黄色

# 对齐
alignment = Alignment(horizontal='center', vertical='center')

# 将样式打包命名
#没有颜色
sty1 = NamedStyle(name='sty1', font=font1,border=border1,alignment=alignment)
#浅黄色
sty2 = NamedStyle(name='sty2', font=font1,border=border1, fill=fill1,alignment=alignment)
#红色
sty3= NamedStyle(name='sty3', font=font1,border=border1, fill=fill2,alignment=alignment)
sty33= NamedStyle(name='sty33', font=font2,border=border1, fill=fill2,alignment=alignment)#字号14
#绿色
sty4= NamedStyle(name='sty4', font=font1,border=border1, fill=fill3,alignment=alignment)
sty44= NamedStyle(name='sty44', font=font2,border=border1, fill=fill3,alignment=alignment)#字号14
#橙色
sty5= NamedStyle(name='sty5', font=font1,border=border1, fill=fill4,alignment=alignment)
sty55= NamedStyle(name='sty55', font=font2,border=border1, fill=fill4,alignment=alignment)#字号14
#蓝色
sty6= NamedStyle(name='sty6', font=font1,border=border1, fill=fill5,alignment=alignment)
sty66= NamedStyle(name='sty66', font=font2,border=border1, fill=fill5,alignment=alignment)#字号14
#粉色
sty7= NamedStyle(name='sty7', font=font1,border=border1, fill=fill6,alignment=alignment)
sty77= NamedStyle(name='sty77', font=font2,border=border1, fill=fill6,alignment=alignment)#字号14
#褐色
sty8= NamedStyle(name='sty8', font=font1,border=border1, fill=fill7,alignment=alignment)
# sty10= NamedStyle(name='sty10', font=font1,border=border1, fill=fill9,alignment=alignment)
#黄色
sty11= NamedStyle(name='sty11', font=font1,border=border1, fill=fill10,alignment=alignment)
#第一行标题
sty12 = NamedStyle(name='sty12', font=font3,border=border1,alignment=alignment)

def get_cal(row_n):
    wb.guess_types=True
    #总成交金额
    total_amount=ws.cell(row=row_n, column=30).value
    #退款率计算  退款率 = 总退款金额 / 总成交金额
    total_refund = ws.cell(row=row_n, column=31).value
    refund_rate=total_refund/total_amount
    ws.cell(row=row_n, column=35, value=refund_rate)
    #商品订单金额占比计算 商品订单金额占比 = 商品订单金额/总成交金额
    order_amount=ws.cell(row=row_n, column=27).value
    order_proportion=order_amount/total_amount
    ws.cell(row=row_n, column=28, value=order_proportion)
    #售卡金额占比计算 售卡金额占比 = 售卡金额 / 总成交金额
    cardsales_amount=ws.cell(row=row_n, column=8).value
    cardsales_proportion=cardsales_amount/total_amount
    ws.cell(row=row_n, column=9, value=cardsales_proportion)
    #升单金额占比计算  升单金额占比 = 升单金额/总成交金额
    upgrade_order_amount=ws.cell(row=row_n, column=12).value
    upgrade_order_proportion=upgrade_order_amount/total_amount
    ws.cell(row=row_n, column=15, value=upgrade_order_proportion)
    #升单率计算  升单率 = 升单人数/新客到店人数
    upgrade_order_num=ws.cell(row=row_n, column=11).value
    new_people_num=ws.cell(row=row_n, column=10).value
    if new_people_num==0:
        upgrade_order_rate=0
    else:
        upgrade_order_rate=np.divide(upgrade_order_num,new_people_num)
    ws.cell(row=row_n, column=14, value=upgrade_order_rate)
    #升单客单价=升单金额/升单人数
    if upgrade_order_num==0:
        upgrade_order_persales=0
    else:
        upgrade_order_persales=np.divide(upgrade_order_amount,upgrade_order_num)
    ws.cell(row=row_n, column=13, value=upgrade_order_persales)
    #复购金额占比计算  复购金额占比 = 复购成交金额/总成交金额
    repurchase_amount=ws.cell(row=row_n, column=18).value
    repurchase_proportion=repurchase_amount/total_amount
    ws.cell(row=row_n, column=21, value=repurchase_proportion)
    #复购客单价=复购成交金额/复购成交人数
    repurchase_done_num=ws.cell(row=row_n, column=17).value
    if repurchase_done_num==0:
        repurchase_persales=0
    else:
        repurchase_persales=np.divide(repurchase_amount,repurchase_done_num)
    ws.cell(row=row_n, column=19, value=repurchase_persales)
    #复购率=复购成交人数/老客到店人数
    old_custermer_num=ws.cell(row=row_n, column=16).value
    if old_custermer_num==0:
        repurchase_rate=0
    else:
        repurchase_rate=np.divide(repurchase_done_num,old_custermer_num)
    ws.cell(row=row_n, column=20, value=repurchase_rate)
    #自拓客金额占比计算 自拓客金额占比=自拓客实收金额/总成交金额
    tuoke_amount=ws.cell(row=row_n, column=24).value
    tuoke_proportion=tuoke_amount/total_amount
    ws.cell(row=row_n, column=25, value=tuoke_proportion)
    #目标完成进度(责任)=总成交金额/责任目标
    duty_target=ws.cell(row=row_n, column=4).value
    if duty_target==0:
        duty_planned_speed=0
    else:
        duty_planned_speed=np.divide(total_amount,duty_target)
    ws.cell(row=row_n, column=36, value=duty_planned_speed)
    #目标完成进度(冲刺)=总成交金额/冲刺业绩
    dash_target=ws.cell(row=row_n, column=5).value
    if dash_target==0:
        dash_planned_speed=0
    else:
        dash_planned_speed=np.divide(total_amount,dash_target)
    ws.cell(row=row_n, column=37, value=dash_planned_speed)
    #目标完成进度(超越)=总成交金额/超越业绩
    overstep_target=ws.cell(row=row_n, column=6).value
    if overstep_target==0:
        overstep_planned_speed=0
    else:
        overstep_planned_speed=np.divide(total_amount,overstep_target)
    ws.cell(row=row_n, column=38, value=overstep_planned_speed)


#总计计算
sum_list=[4,5,6,7,8,10,11,12,16,17,18,22,23,24,26,27,29,30,31,32,33,34]
last_row = ws.max_row+1
for s_column in sum_list:
    total=0
    for f in range(2,last_row):
        one_value = ws.cell(row=f, column=s_column).value
        if one_value is None:
            one_value=0
        total+=one_value
    ws.cell(row=last_row, column=s_column, value=total)
get_cal(last_row)
ws.cell(row=last_row, column=39, value=total_adv_amount)
total_amount1=ws.cell(row=last_row, column=30).value
#投产比
if total_adv_amount==0:
    production_ratio=0
else:
    production_ratio=np.divide(total_amount1,total_adv_amount)
ws.cell(row=last_row, column=40, value=production_ratio)



result=1
for k in range(len(place_list1)):
    result=result+place_dict[place_list1[k]]+1
    #插入一行
    ws.insert_rows(idx=result,amount=1)
    #计算值
    start_index=result-place_dict[place_list1[k]]
    end_index=result
    for sum_column in sum_list:
        res=0
        for index in range(start_index, end_index):
            cellValue = ws.cell(row=index, column=sum_column).value
            res+=cellValue
        ws.cell(row=result, column=sum_column, value=res)
    get_cal(result)
    if place_list1[k]!='平台':
        area_adv_amount=area_amount[place_list1[k]]
        ws.cell(row=result, column=39, value=area_adv_amount)
        total_amount2=ws.cell(row=result, column=30).value
        #投产比
        if area_adv_amount==0:
            production_ratio=0
        else:
            production_ratio=np.divide(total_amount2,area_adv_amount)
        ws.cell(row=result, column=40, value=production_ratio)

ws.insert_rows(idx=0,amount=2)

ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
ws.cell(2, 1).value = '序号'
ws.cell(2, 1).style = sty1

ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)
ws.cell(2, 2).value = '大区'
ws.cell(2, 2).style = sty1

ws.merge_cells(start_row=2, start_column=3, end_row=3, end_column=3)
ws.cell(2, 3).value = '门店名称'
ws.cell(2, 3).style = sty1

ws.merge_cells(start_row=2, start_column=4, end_row=3, end_column=4)
ws.cell(2, 4).value = '责任目标'
ws.cell(2, 4).style = sty11

ws.merge_cells(start_row=2, start_column=5, end_row=3, end_column=5)
ws.cell(2, 5).value = '冲刺目标'
ws.cell(2, 5).style = sty11

ws.merge_cells(start_row=2, start_column=6, end_row=3, end_column=6)
ws.cell(2, 6).value = '超越目标'
ws.cell(2, 6).style = sty11

ws.merge_cells(start_row=2, start_column=7, end_row=2, end_column=9)
ws.cell(2, 7).value = '售卡'
ws.cell(2, 7).style = sty33

ws.merge_cells(start_row=2, start_column=10, end_row=2, end_column=15)
ws.cell(2, 10).value = '新客业绩'
ws.cell(2, 10).style = sty44

ws.merge_cells(start_row=2, start_column=16, end_row=2, end_column=21)
ws.cell(2, 16).value = '老顾客复购业绩'
ws.cell(2, 16).style = sty55

ws.merge_cells(start_row=2, start_column=22, end_row=2, end_column=25)
ws.cell(2, 22).value = '拓客'
ws.cell(2, 22).style = sty66

ws.merge_cells(start_row=2, start_column=26, end_row=2, end_column=28)
ws.cell(2, 26).value = '产品'
ws.cell(2, 26).style = sty77

ws.merge_cells(start_row=2, start_column=29, end_row=3, end_column=29)
ws.cell(2, 29).value = '总成交人数'
ws.cell(2, 29).style = sty8

ws.merge_cells(start_row=2, start_column=30, end_row=3, end_column=30)
ws.cell(2, 30).value = '总成交金额'
ws.cell(2, 30).style = sty8

ws.merge_cells(start_row=2, start_column=31, end_row=3, end_column=31)
ws.cell(2, 31).value = '总退款金额'
ws.cell(2, 31).style = sty8

ws.merge_cells(start_row=2, start_column=32, end_row=3, end_column=32)
ws.cell(2, 32).value = '实际成交金额'
ws.cell(2, 32).style = sty8

ws.merge_cells(start_row=2, start_column=33, end_row=3, end_column=33)
ws.cell(2, 33).value = '服务人数'
ws.cell(2, 33).style = sty8

ws.merge_cells(start_row=2, start_column=34, end_row=3, end_column=34)
ws.cell(2, 34).value = '服务金额'
ws.cell(2, 34).style = sty8

ws.merge_cells(start_row=2, start_column=35, end_row=3, end_column=35)
ws.cell(2, 35).value = '退款率'
ws.cell(2, 35).style = sty8

ws.merge_cells(start_row=2, start_column=36, end_row=3, end_column=36)
ws.cell(2, 36).value = '目标完成进度(责任)'
ws.cell(2, 36).style =  sty11

ws.merge_cells(start_row=2, start_column=37, end_row=3, end_column=37)
ws.cell(2, 37).value = '目标完成进度(冲刺)'
ws.cell(2, 37).style =  sty11

ws.merge_cells(start_row=2, start_column=38, end_row=3, end_column=38)
ws.cell(2, 38).value = '目标完成进度(超越)'
ws.cell(2, 38).style =  sty11

ws.merge_cells(start_row=2, start_column=39, end_row=3, end_column=39)
ws.cell(2, 39).value = '投放金额'
ws.cell(2, 39).style = sty8

ws.merge_cells(start_row=2, start_column=40, end_row=3, end_column=40)
ws.cell(2, 40).value = '投产比'
ws.cell(2, 40).style = sty8


res1=3
for s in range(len(place_list1)):
    res1=res1+place_dict[place_list1[s]]+1
    #合并合计单元格
    ws.merge_cells(start_row=res1, start_column=1, end_row=res1, end_column=3)
    ws.cell(res1, 1).value = '合计'
    ws.cell(res1, 1).style = sty2
    for x in range(4,41):
        ws.cell(res1, x).style = sty2
    start_area=res1-place_dict[place_list1[s]]
    ws.merge_cells(start_row=start_area, start_column=2, end_row=res1-1, end_column=2)
    ws.cell(start_area, 2).value = place_list1[s]
    ws.cell(start_area, 2).style = sty1
final_row = ws.max_row
ws.merge_cells(start_row=final_row, start_column=1, end_row=final_row, end_column=3)
ws.cell(final_row, 1).value = '总计'
ws.cell(final_row, 1).style = sty3
for y in range(4,41):
    ws.cell(final_row, y).style = sty3


ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=40)
ws.cell(1, 1).value = '全国市场门店本月'+date+'累计销售日报表'
ws.cell(1, 1).style = sty12

#产品
ws.cell(3, 26).style = sty7
ws.cell(3, 27).style = sty7
ws.cell(3, 28).style = sty7
#售卡
ws.cell(3, 7).style = sty3
ws.cell(3, 8).style = sty3
ws.cell(3, 9).style = sty3
#新客业绩
ws.cell(3, 10).style = sty4
ws.cell(3, 11).style = sty4
ws.cell(3, 12).style = sty4
ws.cell(3, 13).style = sty4
ws.cell(3, 14).style = sty4
ws.cell(3, 15).style = sty4
#老客户复购业绩
ws.cell(3, 16).style = sty5
ws.cell(3, 17).style = sty5
ws.cell(3, 18).style = sty5
ws.cell(3, 19).style = sty5
ws.cell(3, 20).style = sty5
ws.cell(3, 21).style = sty5
#拓客
ws.cell(3, 22).style = sty6
ws.cell(3, 23).style = sty6
ws.cell(3, 24).style = sty6
ws.cell(3, 25).style = sty6

# 调整行高
ws.row_dimensions[1].height = 31.5
ws.row_dimensions[3].height = 40

# 调整列宽
ws.column_dimensions['C'].width = 42.82
ws.column_dimensions['D'].width = 11.82
ws.column_dimensions['E'].width = 11.82
ws.column_dimensions['F'].width = 11.82
ws.column_dimensions['G'].width = 8.27
ws.column_dimensions['H'].width = 8.75
ws.column_dimensions['I'].width = 11.27
ws.column_dimensions['L'].width = 9.25
ws.column_dimensions['M'].width = 9.38
ws.column_dimensions['O'].width = 12.00
ws.column_dimensions['R'].width = 12.00
ws.column_dimensions['X'].width = 12.48
ws.column_dimensions['Y'].width = 11.00
ws.column_dimensions['AA'].width = 10.13
ws.column_dimensions['AB'].width = 10.13
ws.column_dimensions['AD'].width = 11.00
ws.column_dimensions['AE'].width = 11.00
ws.column_dimensions['AF'].width = 12.48
ws.column_dimensions['AH'].width = 11.00
ws.column_dimensions['AI'].width = 10.38
ws.column_dimensions['AJ'].width = 12.48
ws.column_dimensions['AK'].width = 12.48
ws.column_dimensions['AL'].width = 12.48
ws.column_dimensions['AM'].width = 12.48


#边框设置
def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
s='A1:AN'+str(ws.max_row)
set_border(ws, s)

list_percentage=['I','N','O','T','U','Y','AB','AI','AJ','AK','AL']
for x in list_percentage:
    for y in range(4,ws.max_row+1):
        cell_postition=x+str(y)
        ws[cell_postition].number_format='0.00%'#设置成百分比格式

s_bar1='AJ4:AJ'+str(ws.max_row)
s_bar2='AK4:AK'+str(ws.max_row)
s_bar3='AL4:AL'+str(ws.max_row)

bar_format1 = DataBarRule(start_type='percent', start_value='0', end_type='percent', end_value='100',color="CC3399",showValue="None", minLength=None, maxLength=None)
ws.conditional_formatting.add(s_bar1,bar_format1)
bar_format2 = DataBarRule(start_type='percent', start_value='0', end_type='percent', end_value='100',color="0066FF", showValue="None", minLength=None, maxLength=None)
ws.conditional_formatting.add(s_bar2,bar_format2)
bar_format3 = DataBarRule(start_type='percent', start_value='0', end_type='percent', end_value='100',color="FF0000", showValue="None", minLength=None, maxLength=None)
ws.conditional_formatting.add(s_bar3,bar_format3)

wb.save(filename = './全国市场门店销售报表月度累计'+date+'.xlsx')

def upload_file(file_path, wx_upload_url):
    file_name = file_path.split("/")[-1]
    with open(file_path, 'rb') as f:
        length = os.path.getsize(file_path)
        data = f.read()
    headers = {"Content-Type": "application/octet-stream"}
    params = {
        "filename": file_name,
        "filelength": length,
    }
    file_data = copy(params)
    file_data['file'] = (file_path.split('/')[-1:][0], data)
    encode_data = encode_multipart_formdata(file_data)
    file_data = encode_data[0]
    headers['Content-Type'] = encode_data[1]
    r = requests.post(wx_upload_url, data=file_data, headers=headers)
    media_id = r.json()['media_id']
    return media_id

def qi_ye_wei_xin_file(wx_url, media_id):
    headers = {"Content-Type": "text/plain"}
    data = {
        "msgtype": "file",
        "file": {
            "media_id": media_id
        }
    }
    r = requests.post(
        url=wx_url,
        headers=headers, json=data)

test_report = './全国市场门店销售报表月度累计'+date+'.xlsx'


wx_api_key = 'bf86a1bd-eac3-4018-98cb-19b9df1c08a0'
wx_upload_url = "https://qyapi.weixin.qq.com/cgi-bin/webhook/upload_media?key={}&type=file".format(wx_api_key)
wx_url = 'https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key={}'.format(wx_api_key)
media_id = upload_file(test_report, wx_upload_url)
qi_ye_wei_xin_file(wx_url, media_id)
os.remove('./全国市场门店销售报表月度累计' + date + '.xlsx')
print('删除文件成功')
conn.close()
print('结束时间{}'.format(time.strftime("%Y-%m-%d-%H:%M:%S", time.localtime())))

import pymysql
import pandas as pd
import datetime
import time
import numpy as np
import warnings
import os
from copy import copy
import requests
from urllib3 import encode_multipart_formdata


warnings.filterwarnings("ignore")

x = datetime.date.today()
oneday = datetime.timedelta(days=1)
y = x - oneday
date = str(x)

end_date = str(time.mktime(time.strptime(date, "%Y-%m-%d")))
begin_date = float(end_date) - 86400
begin_date = str(begin_date)

db=pymysql.connect(host='l5vhlstu.ipyingshe.net',database='sys',user='root',password='root123',port=31453,charset='utf8')
sql1='select * from To_the_shop'
df_shops=pd.read_sql(sql1,db)


sql2='select * from area_store'
df_store=pd.read_sql(sql2,db)
df_s=df_store.loc[:,['门店名称']]


st=int(float(begin_date))*1000
ed=int(float(end_date))*1000

df_shop=df_shops.loc[(df_shops['到店时间']<ed)&(df_shops['到店时间']>=st),['到店门店','手机号']]
df_shop['到店门店'] = df_shop['到店门店'].str.strip()

df_lll=df_shop.groupby('到店门店').count().reset_index()
df_lll.rename(columns={df_lll.columns[0]:'门店名称',df_lll.columns[1]:'新客到店人数'},inplace=True)
df_lll['目标业绩']=df_lll['新客到店人数']*3000


sql3='select * from result_day'
df_sales_day=pd.read_sql(sql3,db)


df_sales_day=df_sales_day.loc[df_sales_day['门店名称']!='平台',['门店名称','总成交金额']]
df_sales_day.rename(columns={df_sales_day.columns[1]:'实际业绩'},inplace=True)

df_merge=pd.merge(df_lll,df_sales_day,on='门店名称',how='inner')

df_merge=pd.merge(df_s,df_merge,on='门店名称',how='left')

df_merge.fillna(0,inplace=True)

df_merge['完成率']=df_merge['实际业绩']/df_merge['目标业绩']

df_merge['完成率']=df_merge['完成率'].replace([np.inf, np.nan], 0)

df_merge['完成率'] = df_merge['完成率'].apply(lambda x: format(x, '.2%'))

data_new=dict(zip(df_merge['门店名称'],df_merge['新客到店人数']))
data_target=dict(zip(df_merge['门店名称'],df_merge['目标业绩']))
data_actual=dict(zip(df_merge['门店名称'],df_merge['实际业绩']))
data_donerate=dict(zip(df_merge['门店名称'],df_merge['完成率']))

value_new=df_merge['新客到店人数'].sum()
value_target=df_merge['目标业绩'].sum()
value_actual=df_merge['实际业绩'].sum()
if value_actual==0 or value_target==0:
    value_donerate=0
else:
    value_donerate=value_actual/value_target

#月度累计
str_month = datetime.datetime(y.year, y.month, 1)
str_date = str(str_month)[:10]
begin_date1 = str(time.mktime(time.strptime(str_date, "%Y-%m-%d")))


st1=int(float(begin_date1))*1000


df_shop1=df_shops.loc[(df_shops['到店时间']<ed)&(df_shops['到店时间']>=st1),['到店门店','手机号']]

df_shop1['到店门店'] = df_shop1['到店门店'].str.strip()


df_lll1=df_shop1.groupby('到店门店').count().reset_index()

df_lll1.rename(columns={df_lll1.columns[0]:'门店名称',df_lll1.columns[1]:'新客到店人数'},inplace=True)
df_lll1['目标业绩']=df_lll1['新客到店人数']*3000


sql4='select * from result_month'
df_sales_month=pd.read_sql(sql4,db)

df_sales_month=df_sales_month.loc[df_sales_month['门店名称']!='平台',['门店名称','总成交金额']]
df_sales_month.rename(columns={df_sales_month.columns[1]:'实际业绩'},inplace=True)

df_merge1=pd.merge(df_lll1,df_sales_month,on='门店名称',how='inner')
df_merge1=pd.merge(df_s,df_merge1,on='门店名称',how='left')
df_merge1.fillna(0,inplace=True)

df_merge1['完成率']=df_merge1['实际业绩']/df_merge1['目标业绩']

df_merge1['完成率']=df_merge1['完成率'].replace([np.inf, np.nan], 0)

df_merge1['完成率'] = df_merge1['完成率'].apply(lambda x: format(x, '.2%'))

data_new_all=dict(zip(df_merge1['门店名称'],df_merge1['新客到店人数']))
data_target_all=dict(zip(df_merge1['门店名称'],df_merge1['目标业绩']))
data_actual_all=dict(zip(df_merge1['门店名称'],df_merge1['实际业绩']))
data_donerate_all=dict(zip(df_merge1['门店名称'],df_merge1['完成率']))

value_new_all=df_merge1['新客到店人数'].sum()
value_target_all=df_merge1['目标业绩'].sum()
value_actual_all=df_merge1['实际业绩'].sum()
if value_actual_all==0 or value_target_all==0:
    value_donerate_all=0
else:
    value_donerate_all=value_actual_all/value_target_all


df_al=pd.merge(df_store,df_merge1,on='门店名称',how='left')

df_pers=df_al.loc[:,['市场经理','目标业绩','实际业绩']]
df_pers=df_pers.groupby('市场经理').sum().reset_index()
df_pers['完成率']=df_pers['实际业绩']/df_pers['目标业绩']
df_pers['完成率']=df_pers['完成率'].replace([np.inf, np.nan], 0)
df_pers['完成率'] = df_pers['完成率'].apply(lambda x: format(x, '.2%'))


data_per_target=dict(zip(df_pers['市场经理'],df_pers['目标业绩']))
data_per_actual=dict(zip(df_pers['市场经理'],df_pers['实际业绩']))
data_per_donerate=dict(zip(df_pers['市场经理'],df_pers['完成率']))



df_area=df_al.loc[:,['大区','目标业绩','实际业绩']]
df_area=df_area.groupby('大区').sum().reset_index()
df_area['完成率']=df_area['实际业绩']/df_area['目标业绩']
df_area['完成率']=df_area['完成率'].replace([np.inf, np.nan], 0)
df_area['完成率'] = df_area['完成率'].apply(lambda x: format(x, '.2%'))

data_area_target=dict(zip(df_area['大区'],df_area['目标业绩']))
data_area_actual=dict(zip(df_area['大区'],df_area['实际业绩']))
data_area_donerate=dict(zip(df_area['大区'],df_area['完成率']))

from openpyxl import load_workbook
wb = load_workbook(filename="实时业绩目标表模板.xlsx")
ws = wb.active

prc = ws.columns
s_list = []
prcTuple = tuple(prc)
for idy in range(8, 9):
    for cell in prcTuple[idy]:
        s_list.append(cell.value)
print(s_list)
s_list1 = []
for j in s_list:
    if j not in s_list1 and j !=0 and j !='到店门店' and j is not None:
        s_list1.append(j)
print(s_list1)

for m in s_list1 :
    row1=s_list.index(m)
    ws.cell(row1 + 1, 10).value = data_new_all[m]
    ws.cell(row1 + 1, 11).value = data_target_all[m]
    ws.cell(row1 + 1, 12).value = data_actual_all[m]
    ws.cell(row1 + 1, 13).value = data_donerate_all[m]
    ws.cell(row1 + 1, 14).value = data_new[m]
    ws.cell(row1 + 1, 15).value = data_target[m]
    ws.cell(row1 + 1, 16).value = data_actual[m]
    ws.cell(row1 + 1, 17).value = data_donerate[m]


p_list = []
for idy in range(4, 5):
    for cell in prcTuple[idy]:
        p_list.append(cell.value)
print(p_list)
p_list1 = []
for j in p_list:
    if j not in p_list1 and j !=0 and j !='市场经理' and j is not None:
        p_list1.append(j)
print(p_list1)

for n in p_list1 :
    row2=p_list.index(n)
    ws.cell(row2 + 1, 6).value = data_per_target[n]
    ws.cell(row2 + 1, 7).value = data_per_actual[n]
    ws.cell(row2 + 1, 8).value = data_per_donerate[n]


a_list = []
for idy in range(0, 1):
    for cell in prcTuple[idy]:
        a_list.append(cell.value)
print(a_list)
a_list1 = []
for j in a_list:
    if j not in a_list1 and j !='总计' and j !='大区' and j is not None:
        a_list1.append(j)
print(a_list1)

for t in a_list1 :
    row3=a_list.index(t)
    ws.cell(row3 + 1, 2).value = data_area_target[t]
    ws.cell(row3 + 1, 3).value = data_area_actual[t]
    ws.cell(row3 + 1, 4).value = data_area_donerate[t]

row_last=ws.max_row
ws.cell(row_last, 2).value =value_target_all
ws.cell(row_last, 3).value =value_actual_all
ws.cell(row_last, 4).value =value_donerate_all

ws.cell(row_last, 10).value =value_new_all
ws.cell(row_last, 11).value =value_target_all
ws.cell(row_last, 12).value =value_actual_all
ws.cell(row_last, 13).value =value_donerate_all

ws.cell(row_last, 14).value =value_new
ws.cell(row_last, 15).value =value_target
ws.cell(row_last, 16).value =value_actual
ws.cell(row_last, 17).value =value_donerate

d="%d/%d"%(y.month, y.day)
d1='%d月合计'%(y.month)
ws.cell(1, 10).value =d1
ws.cell(1, 14).value =d

#获取昨日时间
today1=datetime.date.today()
yesterday1=today1-datetime.timedelta(days=1)
yesterday1=str(yesterday1)
wb.save(filename = './实时业绩目标表'+yesterday1+'.xlsx')

def upload_file(file_path, wx_upload_url):
    file_name = file_path.split("/")[-1]
    with open(file_path, 'rb') as f:
        length = os.path.getsize(file_path)
        data = f.read()
    headers = {"Content-Type": "application/octet-stream"}
    params = {
        "filename": file_name,
        "filelength": length,
    }
    file_data = copy(params)
    file_data['file'] = (file_path.split('/')[-1:][0], data)
    encode_data = encode_multipart_formdata(file_data)
    file_data = encode_data[0]
    headers['Content-Type'] = encode_data[1]
    r = requests.post(wx_upload_url, data=file_data, headers=headers)
    media_id = r.json()['media_id']
    return media_id

def qi_ye_wei_xin_file(wx_url, media_id):
    headers = {"Content-Type": "text/plain"}
    data = {
        "msgtype": "file",
        "file": {
            "media_id": media_id
        }
    }
    r = requests.post(
        url=wx_url,
        headers=headers, json=data)

test_report = './实时业绩目标表'+yesterday1+'.xlsx'


wx_api_key = '4bf9ee31-dcf2-4afa-89a9-4fb0c29df9ed'
wx_upload_url = "https://qyapi.weixin.qq.com/cgi-bin/webhook/upload_media?key={}&type=file".format(wx_api_key)
wx_url = 'https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key={}'.format(wx_api_key)
media_id = upload_file(test_report, wx_upload_url)
qi_ye_wei_xin_file(wx_url, media_id)
os.remove('./实时业绩目标表'+yesterday1+'.xlsx')
print('删除文件成功')
print('结束时间{}'.format(time.strftime("%Y-%m-%d-%H:%M:%S", time.localtime())))
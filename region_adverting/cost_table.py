#!/usr/bin/env python
# coding: utf-8

from openpyxl.styles import colors, Font, Fill, NamedStyle
from openpyxl.styles import PatternFill, Border, Side, Alignment
import numpy as np
from openpyxl import load_workbook
import datetime
import requests
from urllib3 import encode_multipart_formdata
import os
from copy import copy

wb = load_workbook(filename="channel_consume.xlsx")
ws = wb.active

today=datetime.date.today() 
oneday=datetime.timedelta(days=1) 
yesterday=today-oneday
date1=str(yesterday)

font1 = Font(name='宋体', size=11, b=True)
font2 = Font(name='宋体', size=11)
line_t = Side(style='thin', color='000000')  # 细边框
border1 = Border(top=line_t, bottom=line_t, left=line_t, right=line_t)

alignment = Alignment(horizontal='center', vertical='center')

fill2=PatternFill('solid', fgColor='66CC66')#绿色

sty1 = NamedStyle(name='sty1', font=font2,border=border1,alignment=alignment)
sty2 = NamedStyle(name='sty2', font=font1,border=border1,alignment=alignment,fill=fill2)

ws.insert_cols(idx=1,amount=3)

ws.cell(1,1).value = '日期'
ws.cell(1,2).value = '投放人'
ws.cell(1,3).value = '账户来源'

ws.merge_cells(start_row=2, start_column=1, end_row=ws.max_row, end_column=1)
ws.cell(2, 1).value = str(yesterday)
ws.cell(2, 1).style = sty1
ws.merge_cells(start_row=2, start_column=2, end_row=ws.max_row, end_column=2)
ws.merge_cells(start_row=2, start_column=3, end_row=ws.max_row, end_column=3)

ws.cell(1,8).value = '总广告费'
ws.cell(1,8).style = sty1
ws.cell(1,9).value='总进线'
ws.cell(1,9).style = sty1
ws.cell(1,10).value='总单粉成本'
ws.cell(1,10).style = sty1

total1=0
ws.merge_cells(start_row=2, start_column=8, end_row=ws.max_row, end_column=8)
for f in range(2,ws.max_row+1):
    value1 = ws.cell(row=f, column=5).value
    if value1 is None:
        value1=0
    total1+=value1
ws.cell(row=2, column=8, value=total1)
ws.cell(row=2, column=8).style = sty1

total2=0
ws.merge_cells(start_row=2, start_column=9, end_row=ws.max_row, end_column=9)
for f1 in range(2,ws.max_row+1):
    value2 = ws.cell(row=f1, column=6).value
    if value2 is None:
        value2=0
    total2+=value2
ws.cell(row=2, column=9, value=total2)
ws.cell(row=2, column=9).style = sty1

ws.merge_cells(start_row=2, start_column=10, end_row=ws.max_row, end_column=10)
all_adv_cost=ws.cell(row=2, column=8).value    
all_inline=ws.cell(row=2, column=9).value
if all_inline==0:
    single_cost=0
else:
    single_cost=np.divide(all_adv_cost,all_inline)      
ws.cell(row=2, column=10, value=single_cost)
ws.cell(row=2, column=10).style = sty1     

for y in range(1,11):
    ws.cell(1, y).style = sty2

ws.column_dimensions['A'].width = 13.75

for col in  ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in col:
            cell.alignment = Alignment(horizontal='center', vertical='center')

#边框设置
def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
s='A1:J'+str(ws.max_row)
set_border(ws, s) 

wb.save(filename = f"医普茂{date1}成本表-消费.xlsx")

wb1 = load_workbook(filename="city_channel.xlsx")
ws1 = wb1.active

prj = ws1.columns
place_list=[]
prjTuple = tuple(prj)  
for idx in range(0,1):# 遍历A、B、C、D列  #(0,len(prjTuple)) 遍历全部列
    for cell in prjTuple[idx]:
        if (cell.value !='地区' and cell.value is not None):
#         print(cell.coordinate, cell.value)
            place_list.append(cell.value)


place_dict= {}
for i in place_list:
    place_dict[i] = place_list.count(i)


place_list1=[]
for j in place_list:
    if j not in place_list1:
        place_list1.append(j)


ws1.insert_cols(idx=1,amount=1)
ws1.cell(1,1).value = '日期'

ws1.merge_cells(start_row=2, start_column=1, end_row=ws1.max_row, end_column=1)
ws1.cell(2, 1).value = str(yesterday)
ws1.cell(2, 1).style = sty1

column_list=[2,7,8,9]
for d in column_list:
    for place in place_list1:
        a=place_list.index(place)+2
        b=place_dict[place]-1
        if b!=0:
            c=ws1.cell(row=a, column=d).value
            ws1.merge_cells(start_row=a, start_column=d, end_row=a+b, end_column=d)
            ws1.cell(a, d).value = c
            ws1.cell(a, d).style = sty1

for y in range(1,10):
    ws1.cell(1, y).style = sty2

ws1.column_dimensions['A'].width = 13.75

for col in  ws1.iter_cols(min_row=2, max_row=ws1.max_row, min_col=1, max_col=ws1.max_column):
        for cell in col:
            cell.alignment = Alignment(horizontal='center', vertical='center')

#边框设置
s1='A1:I'+str(ws1.max_row)
set_border(ws1, s1) 

wb1.save(filename = f"医普茂{date1}成本表-地区.xlsx")

def upload_file(file_path, wx_upload_url):
    file_name = file_path.split("/")[-1]
    with open(file_path, 'rb') as f:
        length = os.path.getsize(file_path)
        data = f.read()
    headers = {"Content-Type": "application/octet-stream"}
    params = {
        "filename": file_name,
        "filelength": length,
    }
    file_data = copy(params)
    file_data['file'] = (file_path.split('/')[-1:][0], data)
    encode_data = encode_multipart_formdata(file_data)
    file_data = encode_data[0]
    headers['Content-Type'] = encode_data[1]
    r = requests.post(wx_upload_url, data=file_data, headers=headers)
    media_id = r.json()['media_id']
    return media_id

def qi_ye_wei_xin_file(wx_url, media_id):
    headers = {"Content-Type": "text/plain"}
    data = {
        "msgtype": "file",
        "file": {
            "media_id": media_id
        }
    }
    r = requests.post(
        url=wx_url,
        headers=headers, json=data)

test_report = f"医普茂{date1}成本表-消费.xlsx"
test_report1=f"医普茂{date1}成本表-地区.xlsx"

wx_api_key = 'cd68ed22-725e-4b4e-b858-4d5813c554c9'
wx_upload_url = "https://qyapi.weixin.qq.com/cgi-bin/webhook/upload_media?key={}&type=file".format(wx_api_key)
wx_url = 'https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key={}'.format(wx_api_key)
media_id = upload_file(test_report, wx_upload_url)
qi_ye_wei_xin_file(wx_url, media_id)

media_id1 = upload_file(test_report1, wx_upload_url)
qi_ye_wei_xin_file(wx_url, media_id1)

os.remove(f"医普茂{date1}成本表-消费.xlsx")
os.remove(f"医普茂{date1}成本表-地区.xlsx")


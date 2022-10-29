#!/usr/bin/env python
# coding: utf-8

import glob
import re
import pandas as pd
import pymongo
import pymysql
import time
from functools import reduce
from openpyxl.styles import colors, Font, Fill, NamedStyle
from openpyxl.styles import PatternFill, Border, Side, Alignment
import numpy as np
import datetime
import os
from copy import copy
import requests
from urllib3 import encode_multipart_formdata
import warnings

warnings.filterwarnings("ignore")

today = datetime.date.today()
oneday = datetime.timedelta(days=1)
yesterday = today - oneday
now_time = str(yesterday)

# filenames1 = glob.glob('D:/下载数据/*.xls')
# dfs1 = [pd.read_excel(i) for i in filenames1]
# df2 = pd.concat(dfs1, sort=False)
#
# df2.drop_duplicates(inplace=True)

# df2.to_excel(f'消耗汇总({now_time}).xlsx',index=False)
df_cc=pd.read_excel('消耗.xlsx')
df_cc=df_cc.loc[:,['advertiser_id','cost']]

df_inn=pd.read_excel('account_info.xlsx')

df_inn=df_inn.loc[:,['广告主ID','用户名']]
df_inn.rename(columns={df_inn.columns[1]:'账户',df_inn.columns[0]:'advertiser_id'},inplace=True)
df_inn.drop_duplicates(subset=None, keep='last', inplace=True)
df=pd.merge(df_cc,df_inn,on='advertiser_id',how='left')
df=df.loc[:,['账户','cost']]
df.rename(columns={df.columns[1]:'消耗'},inplace=True)


df_clue = pd.read_excel(f'导线数据(0-24).xlsx')

df_clue=df_clue.loc[:,['telephone','location','advertiser_name']]

df_clue.rename(columns={df_clue.columns[0]:'电话',df_clue.columns[1]:'自动定位城市',df_clue.columns[2]:'广告主名称'},inplace=True)
# df_clue = df_clue.loc[:, ['电话', '自动定位城市', '广告主名称']]

import copy

df_clue_copy = df_clue.copy()

df_clue_copy = df_clue_copy.loc[:, ['自动定位城市', '广告主名称']]

df_clue_copy.drop_duplicates(inplace=True)

df_clue_copy.dropna(subset=['自动定位城市'], inplace=True)

df_clue_copy.drop_duplicates(subset=['广告主名称'], keep=False, inplace=True)

loc_dict = df_clue_copy.set_index(['广告主名称'])['自动定位城市'].to_dict()

dict_keys = list(loc_dict.keys())
dict_values = list(loc_dict.values())
for x in range(len(dict_keys)):
    df_clue.loc[(df_clue['广告主名称'] == dict_keys[x]) & (df_clue['自动定位城市']).isnull(), '自动定位城市'] = dict_values[x]

df_clue.insert(3, '市', '')


# 市
def extract_city(x):
    if str(x) != 'nan':
        y = str(x).split('+')
        return y[1]


df_clue['市'] = df_clue['自动定位城市'].apply(lambda x: extract_city(x))

df_phone = df_clue.loc[:, ['电话', '市']]

df_phone.rename(columns={df_phone.columns[0]: 'phone'}, inplace=True)

df_phone = df_phone.astype({'phone': 'str'})

df_clue = df_clue.loc[:, ['市', '广告主名称']]

df_clue.rename(columns={df_clue.columns[1]: '用户名'}, inplace=True)

df_clue = df_clue.value_counts().reset_index(name='数量')

df_clue1 = df_clue.loc[:, ['用户名', '数量']]
df_clue1 = df_clue1.groupby('用户名').sum().reset_index()

dfx = pd.read_excel('account_info.xlsx')

dfx1 = dfx.loc[:, ['用户名', '投放地域']]

dfx1['市'] = dfx1['投放地域'].apply(lambda x: re.findall(r'[\u4e00-\u9fa5]+', x)[0])

df.rename(columns={df.columns[0]: '用户名'}, inplace=True)

df_merge1 = pd.merge(df, df_clue1, on='用户名', how='left')

df_merge1['数量'].fillna(0, inplace=True)

df_cut = df_merge1.loc[(df_merge1['数量'] == 0.0) & (df_merge1['消耗'] != 0.00), :]

df_cut1 = pd.merge(df_cut, dfx1, on='用户名', how='left')

df_cut2 = df_cut1.loc[:, ['用户名']]

df_cut3 = df_cut2.value_counts().reset_index(name='次数')

df_cut4 = pd.merge(df_cut1, df_cut3, on='用户名', how='left')

df_cut4['消耗1'] = df_cut4['消耗'] / df_cut4['次数']

df_clue_copy2 = df_clue.copy()
df_clue_copy2 = df_clue_copy2.loc[:, ['市', '用户名']]
df_clue_copy2.drop_duplicates(subset=None, keep=False, inplace=True)

df_cut5 = df_cut4.loc[:, ['市', '投放地域', '消耗1', '数量']]

df_cut5.insert(4, '单粉成本', 0)

df_cut5.rename(columns={df_cut5.columns[2]: '广告费'}, inplace=True)

df_none = df_merge1.loc[(df_merge1['消耗'] == 0.00) & (df_merge1['数量'] != 0.0), :]

df_none1 = pd.merge(dfx1, df_none, on='用户名', how='inner')

df_none1['单粉成本'] = 0.0

df_none2 = pd.merge(df_none1, df_clue_copy2, on=['市', '用户名'], how='inner')

df_none2 = df_none2.loc[:, ['市', '投放地域', '消耗', '数量', '单粉成本']]

df_none2.rename({df_none2.columns[2]: '广告费'}, inplace=True)

df_merge1.eval("""单粉成本=消耗/数量""", inplace=True)

df_merge1 = df_merge1.loc[:, ['用户名', '单粉成本']]

df_last = pd.merge(df_clue, df_merge1, on='用户名', how='left')

df_last['广告费'] = df_last['单粉成本'] * df_last['数量']

df_last1 = pd.merge(df_last, dfx1, on=['用户名', '市'], how='left')

df_last1 = df_last1.loc[:, ['市', '投放地域', '广告费', '数量', '单粉成本']]

df_last2 = pd.concat([df_last1, df_cut5, df_none2])

df_last2 = df_last2.sort_values(by="市", ascending=False)

df_last2['广告费'].fillna(0, inplace=True)
df_last2.drop_duplicates(subset=['投放地域', '数量'], inplace=True)

df_channel = df_last2.loc[:, ['投放地域', '广告费', '数量', '单粉成本']]

df_channel.rename(columns={df_channel.columns[0]: '渠道名', df_channel.columns[2]: '表单进线'}, inplace=True)

df_channel.to_excel('channel_consume.xlsx', index=False)

df_channel2 = df_last2.loc[:, ['市', '投放地域', '广告费', '数量', '单粉成本']]
df_channel2.rename(
    columns={df_channel2.columns[0]: '地区', df_channel2.columns[1]: '文案编号', df_channel2.columns[3]: '表单进线'},
    inplace=True)

df_city = df_last2.loc[:, ['市', '广告费', '数量']]

df_city = df_city.groupby('市').sum().reset_index()

df_city['进线成本'] = df_city['广告费'] / df_city['数量']

df_city = df_city.loc[:, ['市', '广告费', '进线成本', '数量']]

df_city.rename(columns={df_city.columns[0]: '城市', df_city.columns[1]: '广告总额', df_city.columns[3]: '进线人数'}, inplace=True)

df_city1 = df_city.loc[:, ['城市', '广告总额', '进线人数', '进线成本']]
df_city1.rename(columns={df_city1.columns[0]: '地区', df_city1.columns[1]: '总广告费', df_city1.columns[2]: '总进线',
                         df_city1.columns[3]: '总单粉成本'}, inplace=True)

df_city_channel = pd.merge(df_channel2, df_city1, on='地区', how='inner')
df_city_channel2 = df_city_channel.sort_values(by="地区", ascending=False)
df_city_channel2.to_excel('city_channel.xlsx', index=False)

df_city.to_excel('city_consume.xlsx', index=False)

# # 删除文件
# def del_file(path):
#     for j in os.listdir(path):
#         file_data = path + "\\" + j
#         if  j.endswith('.xls'):
#             os.remove(file_data)
# del_file('D:/下载数据/')


today1 = datetime.date.today()
# 昨天日期
yesterday1 = today1 - datetime.timedelta(days=1)
# 昨天开始时间戳
begin_date1 = str(time.mktime(time.strptime(str(yesterday1), "%Y-%m-%d")))

conn = pymongo.MongoClient('mongodb://{}:{}@{}:{}/?authSource={}'
                           .format("cda_epmall", "3PUkMsWCTu424Renrffb",
                                   "dds-wz99d33fe2f72610a042-pub.mongodb.rds.aliyuncs.com",
                                   "3717", "epmall"))
conn_epmall = conn["epmall"]

col_ccc_user = conn_epmall['ccc_user']  # 小程序用户表
df_ccc_user = pd.DataFrame(list(col_ccc_user.find({'create_time': {'$lt': begin_date1}}, {'_id': 0, 'phone': 1})))

df_ccc_user = df_ccc_user.astype({'phone': 'str'})
df_duplicate = pd.merge(df_phone, df_ccc_user, on='phone', how='inner')

df_duplicate1 = df_duplicate.copy()
df_duplicate1.drop_duplicates(inplace=True)
df_duplicate1 = df_duplicate1.groupby('市').count().reset_index()
df_duplicate1.rename(columns={df_duplicate1.columns[0]: '城市', df_duplicate1.columns[1]: '系统重粉数'}, inplace=True)
df_data11 = pd.merge(df_city, df_duplicate1, on='城市', how='left')

df_duplicate = df_duplicate.groupby('市').count().reset_index()
df_duplicate.rename(columns={df_duplicate.columns[0]: '城市', df_duplicate.columns[1]: '重粉数'}, inplace=True)

df_data = pd.merge(df_data11, df_duplicate, on='城市', how='left')

df_data['表单重粉数'] = df_data['重粉数'] - df_data['系统重粉数']

df_data = df_data.loc[:, ['城市', '广告总额', '进线成本', '进线人数', '系统重粉数', '表单重粉数', '重粉数']]

df_data['重粉率'] = df_data['重粉数'] / df_data['进线人数']

df_city_province = pd.read_excel('省份城市表.xlsx')

df_middle_result = pd.merge(df_city_province, df_data, on='城市', how='left')

# 今天日期
today = datetime.date.today()
# 昨天日期
yesterday = today - datetime.timedelta(days=1)
# 昨天开始时间戳
begin_date = str(time.mktime(time.strptime(str(yesterday), "%Y-%m-%d")))
# 昨天结束时间戳
end_date = str(time.mktime(time.strptime(str(today), "%Y-%m-%d")))

print('开始时间{}'.format(time.strftime("%Y-%m-%d-%H:%M:%S", time.localtime())))
# 连接mongdb
conn = pymongo.MongoClient('mongodb://{}:{}@{}:{}/?authSource={}'
                           .format("cda_epmall", "3PUkMsWCTu424Renrffb",
                                   "dds-wz99d33fe2f72610a042-pub.mongodb.rds.aliyuncs.com",
                                   "3717", "epmall"))
conn_epmall = conn["epmall"]

df_order_all = conn_epmall['epmall_orders_100063']  # 全量订单表

df_order_all = pd.DataFrame(list(df_order_all.find({'state': {'$gte': 1, '$lte': 9}},
                                                   {'_id': 0, 'state': 1, 'order_no': 1, 'create_date': 1, 'user_id': 1,
                                                    'nick': 1, 'mobile': 1,
                                                    'isback': 1, 'commodity_id': 1, 'commodity_name': 1,
                                                    'order_type': 1,
                                                    'recommend_store_id': 1, 'third_platform': 1, 'is_selfpick': 1,
                                                    'pick_store_id': 1, 'amount_total': 1, 'totalDiscount': 1,
                                                    'recharge_balance': 1})))
df_order_all['recommend_store_id'] = df_order_all['recommend_store_id'].mask(df_order_all['order_type'] == 0,
                                                                             df_order_all['pick_store_id'])
df_order_all.replace('', '平台', inplace=True)
df_order_all['recommend_store_id'].fillna('平台', inplace=True)

# In[279]:


df_order_all0 = conn_epmall['epmall_orders_100063']  # 全量订单表 退款找订单id用

df_order_all0 = pd.DataFrame(list(df_order_all0.find({'state': {'$gte': 1, '$lte': 9}},
                                                     {'_id': 0, 'state': 1, 'order_no': 1, 'create_date': 1,
                                                      'user_id': 1, 'nick': 1, 'mobile': 1,
                                                      'isback': 1, 'commodity_id': 1, 'commodity_name': 1,
                                                      'order_type': 1,
                                                      'recommend_store_id': 1, 'third_platform': 1, 'is_selfpick': 1,
                                                      'pick_store_id': 1, 'amount_total': 1, 'totalDiscount': 1,
                                                      'recharge_balance': 1})))
df_order_all0['recommend_store_id'] = df_order_all0['recommend_store_id'].mask(df_order_all0['order_type'] == 0,
                                                                               df_order_all0['pick_store_id'])
df_order_all0.replace('', '平台', inplace=True)
df_order_all0['recommend_store_id'].fillna('平台', inplace=True)

df_recharge_all = conn_epmall['epmall_recharge_log_100063']  # 全量充值表
df_recharge_all = pd.DataFrame(
    list(df_recharge_all.find({'pay_state': 1}, {'_id': 0, 'activity_store_id': 1, 'deal_no': 1})))
df_recharge_all.replace('', '平台', inplace=True)
df_recharge_all['activity_store_id'].fillna('平台', inplace=True)

x = datetime.date.today()
oneday = datetime.timedelta(days=1)
y = x - oneday
str_month = datetime.datetime(y.year, y.month, 1)
str_date = str(str_month)[:10]
en_date = str(y)
date = str(x)

begin_date = str(time.mktime(time.strptime(str_date, "%Y-%m-%d")))

end_date = str(time.mktime(time.strptime(date, "%Y-%m-%d")))

date_list = []
while True:
    date_list.append(str(x))
    if str(x) == str(str_month)[:10]:
        break
    x = x - oneday

begin_date = float(end_date) - 86400
begin_date = str(begin_date)
print(begin_date)
# In[282]:


col_order = conn_epmall['epmall_orders_100063']  # 订单表

df_order = pd.DataFrame(
    list(col_order.find({'create_date': {'$gt': begin_date, '$lt': end_date}, 'state': {'$gte': 1, '$lte': 9}},
                        {'_id': 0, 'state': 1, 'order_no': 1, 'create_date': 1, 'user_id': 1, 'nick': 1, 'mobile': 1,
                         'isback': 1, 'commodity_id': 1, 'commodity_name': 1, 'order_type': 1,
                         'recommend_store_id': 1, 'third_platform': 1, 'is_selfpick': 1, 'pick_store_id': 1,
                         'amount_total': 1, 'totalDiscount': 1, 'recharge_balance': 1})))
df_order.eval(""" accounted_money= amount_total - totalDiscount - recharge_balance""", inplace=True)
df_order['accounted_money'] = df_order['accounted_money'] / 100
df_order['recommend_store_id'] = df_order['recommend_store_id'].mask(df_order['order_type'] == 0,
                                                                     df_order['pick_store_id'])
df_order.replace('', '平台', inplace=True)
df_order['recommend_store_id'].fillna('平台', inplace=True)
df_order = df_order[df_order['accounted_money'] > 0]

col_app_user = conn_epmall['epmall_app_user_100063']  # 小程序用户表
df_app_user = pd.DataFrame(list(col_app_user.find({},
                                                  {'_id': 0, 'user_id': 1, 'promote_orders': 1,
                                                   'promote_orders_time': 1, 'repurchase_orders': 1,
                                                   'repurchase_orders_time': 1, 'channel_belong_name': 1,
                                                   'deliver_id': 1})))  # 用户id,是否升单,升单时间,是否复购,复购时间

col_serve = conn_epmall['epmall_service_record_100063']  # 服务记录表
df_serve = pd.DataFrame(list(col_serve.find({'create_time': {'$gt': begin_date, '$lt': end_date}},
                                            {'_id': 0, 'user_id': 1, 'store_id': 1, 'creat_time': 1,
                                             'afterDiscount': 1})))
df_serve['store_id'].fillna('平台', inplace=True)
df_serve['afterDiscount'] = df_serve['afterDiscount'] / 100

col_recharge = conn_epmall['epmall_recharge_log_100063']  # 充值表
df_recharge = pd.DataFrame(list(
    col_recharge.find({'recharge_date': {'$gt': begin_date, '$lt': end_date}, 'pay_state': 1},
                      {'_id': 0, 'user_id': 1, 'amount_pay': 1, 'activity_store_id': 1, 'deal_no': 1,
                       'recharge_date': 1})))

df_recharge['amount_pay'] = df_recharge['amount_pay'] / 100
if 'activity_store_id' in df_recharge.columns:
    df_recharge['activity_store_id'].fillna('平台', inplace=True)
else:
    df_recharge['activity_store_id'] = '平台'

df_recharge['user_id'] = pd.to_numeric(df_recharge['user_id'], errors='raise')

col_refund = conn_epmall['epmall_refund_100063']  # 充值退款表  'create_time':{'$gt':begin_date,'$lt':end_date},
df_refund = pd.DataFrame(list(col_refund.find({'refund_time': {'$gt': begin_date, '$lt': end_date}, 'state': 2},
                                              {'_id': 0, 'user_id': 1, 'refund_money': 1, 'deal_no': 1})))
if len(df_refund) != 0:
    df_refund['refund_money'] = df_refund['refund_money'] / 100
    # df_refund.to_excel('退款充值.xlsx')
    df_refund = df_refund.rename(columns={'refund_money': 'amount_refund'})

col_back = conn_epmall['epmall_back_order']  # 退单表
df_back = pd.DataFrame(list(col_back.find({'cashier_data': {'$gt': begin_date, '$lt': end_date}, 'back_result': 4},
                                          {'_id': 0, 'order_no': 1, 'user_id': 1, 'recharge_balance': 1,
                                           'commodity_id': 1, 'amount_refund': 1})))
if len(df_back) != 0:
    df_back['wxrefund_money'] = df_back['amount_refund'] - df_back['recharge_balance']
    df_back['wxrefund_money'] = df_back['wxrefund_money'] / 100
    df_back = df_back[df_back['wxrefund_money'] > 0]

# ### 卖卡数量 卖卡金额


col_commodity = conn_epmall['epmall_commodity_100063']  # 商品表
df_commodity = pd.DataFrame(list(col_commodity.find({},
                                                    {'_id': 0, 'commodity_id': 1,
                                                     'variety_id': 1})))

variety_id = ['1646038407932', '1646209380573', '1646491684554', '1650421003620']
df_commodity = df_commodity[df_commodity['variety_id'].isin(variety_id)]
commodity_id_list = list(df_commodity['commodity_id'])  # 卡商品id

df_order_all = df_order_all[df_order_all['commodity_id'].isin(commodity_id_list)]  # 此为全量售卡订单

order_card = df_order[df_order['commodity_id'].isin(commodity_id_list)]
# 此为售卡订单
len(order_card)

# 算人数 来自app订单和来自微信小程序要去重
third_platform1 = [0, 3, 5, 6]
order_card1 = order_card[order_card['third_platform'].isin(third_platform1)].drop_duplicates(subset=['user_id'])
len(order_card1)

# 拼接
third_platform2 = [1, 2, 4]
order_card2 = order_card[order_card['third_platform'].isin(third_platform2)]
order_card_new = pd.concat([order_card1, order_card2])
# order_card_new.to_excel('卖卡人数.xlsx')

## 卖卡人数 卖卡金额

card_sum = order_card_new.groupby('recommend_store_id')['user_id'].count().reset_index()

card_sum = card_sum.rename(columns={'user_id': '售卡人数'})
card_sum.sum()

card_money = order_card_new.groupby('recommend_store_id')['accounted_money'].sum().reset_index()

card_money = card_money.rename(columns={'accounted_money': '售卡金额'})

# 计算升单金额及人数、卡升单金额及人数、复购金额及人数时，需增加体验卡剔除的机制。
df_order_remove = df_order[~df_order['commodity_id'].isin(commodity_id_list)]

# ### 升单人数 ,升单金额

# 升单人数 (有app_user 表里 promote_orders =1 都算) 当月
# 升单金额(有app_user 表里 promote_orders =1 ,升单订单的筛选方法:promote_orders_time<=订单时间<repurchase_orders_time)
# up['user_id'].unique()

col_user_up = conn_epmall['epmall_app_user_100063']  # 小程序用户表promote_orders_time 为当月的
df_user_up = pd.DataFrame(list(col_user_up.find({'promote_orders_time': {'$gt': begin_date, '$lt': end_date}},
                                                {'_id': 0, 'user_id': 1, 'promote_orders': 1, 'promote_orders_time': 1,
                                                 'repurchase_orders': 1,
                                                 'repurchase_orders_time': 1})))

df_promote_orders = df_user_up[df_user_up['promote_orders'] == 1]

# promote_orders=1者

order_up = pd.merge(df_promote_orders, df_order, on='user_id')
recharge_up = pd.merge(df_promote_orders, df_recharge, on='user_id')

recharge_up = recharge_up.loc[:, ['activity_store_id', 'user_id']]
recharge_up.rename(columns={'activity_store_id': 'recommend_store_id'}, inplace=True)

up_sum = order_up[order_up['accounted_money'] != 0]
up_sum = up_sum.loc[:, ['recommend_store_id', 'user_id']]
up_sum = pd.concat([up_sum, recharge_up])
up_sum = up_sum.drop_duplicates(subset='user_id')

up_sum = up_sum.groupby('recommend_store_id')['user_id'].count().reset_index()  # 升单人数
up_sum = up_sum.rename(columns={'user_id': '升单人数'})

df_user_up1 = pd.merge(df_user_up, df_order_remove, on='user_id', how='inner')  # 合并订单和user
df_user_up2 = pd.merge(df_user_up, df_recharge, on='user_id', how='inner')  # 合并充值和user

# 计算升单订单金额
df_user_promote = df_user_up1[df_user_up1['repurchase_orders'] == 1]
df_user_promote = df_user_promote[(df_user_promote['create_date'] < df_user_promote['repurchase_orders_time']) & (
        df_user_promote['promote_orders_time'] <= df_user_promote['create_date'])]

df_user_promote0 = df_user_up1[df_user_up1['repurchase_orders'] == 0]
df_user_promote0 = df_user_promote0[(df_user_promote0['promote_orders_time'] <= df_user_promote0['create_date'])]

df_user_promote0 = pd.concat([df_user_promote0, df_user_promote])

df_user_promote0 = df_user_promote0.loc[:, ['user_id', 'accounted_money', 'recommend_store_id']]

# 计算充值金额
df_user_repurchase = df_user_up2[df_user_up2['repurchase_orders'] == 1]
df_user_repurchase = df_user_repurchase[
    (df_user_repurchase['recharge_date'] < df_user_repurchase['repurchase_orders_time']) & (
            df_user_repurchase['promote_orders_time'] <= df_user_repurchase['recharge_date'])]

df_user_repurchase0 = df_user_up2[df_user_up2['repurchase_orders'] == 0]
df_user_repurchase0 = df_user_repurchase0[
    (df_user_repurchase0['promote_orders_time'] <= df_user_repurchase0['recharge_date'])]
df_user_repurchase0 = pd.concat([df_user_repurchase0, df_user_repurchase])

df_user_repurchase0.rename(columns={'activity_store_id': 'recommend_store_id', 'amount_pay': 'accounted_money'},
                           inplace=True)

df_user_repurchase0 = df_user_repurchase0.loc[:, ['user_id', 'accounted_money', 'recommend_store_id']]

up_money = pd.concat([df_user_repurchase0, df_user_promote0])

up_money = up_money.groupby('recommend_store_id')['accounted_money'].sum().reset_index()
up_money = up_money.rename(columns={'accounted_money': '升单金额'})  # 计算升单金额

# ### 卡到店人数,卡升单人数,卡升单金额,卡升单转化率 ,卡升单客单价

# 卡到店人数 :先筛选出本月购卡者 然后与服务记录表相关联订单
#
# 卡升单人数 先筛选出本月购卡者 查找 app_user 表里 promote_orders =1 且promote_orders_time 在购卡订单时间之后者
#
# 卡升单金额 卡升单客户的升单订单实付金额 (升单订单的筛选方法:promote_orders_time<=订单时间<repurchase_orders_time)
# 卡升单转化率 =卡升单人数/卡到店人数
#
# 卡升单客单价 =卡升单金额/卡升单人数

# 卡到店人数
card_shop = pd.merge(df_order_all, df_serve, on='user_id', how='left')
card_shop = card_shop.drop_duplicates('user_id')
# card_shop.to_excel('卡到店人数.xlsx')
card_shop_sum = card_shop.groupby('store_id')['user_id'].count().reset_index()

card_shop_sum = card_shop_sum.rename(columns={'user_id': '卡到店人数'})
card_shop_sum.rename(columns={'store_id': 'recommend_store_id'}, inplace=True)

card_user_up = conn_epmall['epmall_app_user_100063']  # 小程序用户表promote_orders_time 为当月的
card_user_up = pd.DataFrame(list(card_user_up.find({'promote_orders_time': {'$gt': begin_date, '$lt': end_date}},
                                                   {'_id': 0, 'user_id': 1, 'promote_orders': 1,
                                                    'promote_orders_time': 1, 'repurchase_orders': 1,
                                                    'repurchase_orders_time': 1})))

# 卡升单人数  当月购卡并且当月升单


# ### 卡升单金额

card_up_sum2 = pd.merge(df_order_all['user_id'], card_user_up, on='user_id', how='inner')
card_up_sum2 = pd.merge(card_up_sum2, df_order, on='user_id', how='inner')
card_up_sum2 = card_up_sum2[~card_up_sum2['order_no'].isin(list(order_card['order_no']))]
card_up_sum2 = card_up_sum2.drop_duplicates(subset='order_no')

card_up_sum3 = pd.merge(df_order_all['user_id'], card_user_up, on='user_id', how='inner')
card_up_sum3 = pd.merge(card_up_sum3, df_recharge, on='user_id', how='inner')
card_up_sum3 = card_up_sum3.drop_duplicates(subset='deal_no')

# #计算卡升单订单金额
df_up_orders = card_up_sum2[card_up_sum2['repurchase_orders'] == 1]
df_up_orders = df_up_orders[(df_up_orders['create_date'] < df_up_orders['repurchase_orders_time']) & (
        df_up_orders['promote_orders_time'] <= df_up_orders['create_date'])]

df_up_orders0 = card_up_sum2[card_up_sum2['repurchase_orders'] == 0]
df_up_orders0 = df_up_orders0[(df_up_orders0['promote_orders_time'] <= df_up_orders0['create_date'])]

df_up_orders0 = pd.concat([df_up_orders0, df_up_orders0])

df_up_orders0 = df_up_orders0.drop_duplicates(subset='order_no')
# df_up_orders0.to_excel('卡升订单.xlsx')
df_up_orders0 = df_up_orders0.loc[:, ['user_id', 'accounted_money', 'recommend_store_id']]

# 计算卡升单充值金额
df_up_recharge = card_up_sum3[card_up_sum3['repurchase_orders'] == 1]
df_up_recharge = df_up_recharge[(df_up_recharge['recharge_date'] <
                                 df_up_recharge['repurchase_orders_time']) & (df_up_recharge['promote_orders_time']
                                                                              <= df_up_recharge['recharge_date'])]

df_up_recharge0 = card_up_sum3[card_up_sum3['repurchase_orders'] == 0]
df_up_recharge0 = df_up_recharge0[(df_up_recharge0['promote_orders_time'] <= df_up_recharge0['recharge_date'])]
df_up_recharge0 = pd.concat([df_up_recharge0, df_up_recharge])
# df_up_recharge0.to_excel('卡升充值.xlsx')
df_up_recharge0.rename(columns={'activity_store_id': 'recommend_store_id', 'amount_pay': 'accounted_money'},
                       inplace=True)

df_up_recharge0 = df_up_recharge0.loc[:, ['user_id', 'accounted_money', 'recommend_store_id']]

card_up_money = pd.concat([df_up_recharge0, df_up_orders0])

card_up_money1 = card_up_money.groupby('recommend_store_id')['accounted_money'].sum().reset_index()
card_up_money1 = card_up_money1.rename(columns={'accounted_money': '卡升单金额'})

# 卡升单金额


# ### 卡升单人数

card_up_sum = card_up_money.drop_duplicates('user_id')
card_up_sum = card_up_sum.groupby('recommend_store_id')['user_id'].count().reset_index()
card_up_sum = card_up_sum.rename(columns={'user_id': '卡升单人数'})
card_up_sum.head(50)

# ### 服务人数,服务金额

# 服务人数
serve_sum = df_serve.drop_duplicates('user_id')
serve_sum = serve_sum.groupby('store_id')['user_id'].count().reset_index()
serve_sum.rename(columns={'user_id': '服务人数', 'store_id': 'recommend_store_id'}, inplace=True)

# 服务金额
serve_money = df_serve.groupby('store_id')['afterDiscount'].sum().reset_index()
serve_money.rename(columns={'afterDiscount': '服务金额', 'store_id': 'recommend_store_id'}, inplace=True)


# ### 老客到店人数,复购成交人数,复购金额

def f(x):
    return datetime.datetime.fromtimestamp(int(x)).replace(hour=0, minute=0, second=0, microsecond=0)


# 老客到店人数
col_serve1 = conn_epmall['epmall_service_record_100063']  # 服务记录表
df_serve1 = pd.DataFrame(list(col_serve1.find({'create_time': {'$gt': begin_date, '$lt': end_date}},
                                              {'_id': 0, 'user_id': 1, 'store_id': 1, 'create_time': 1})))
df_serve1['store_id'].fillna('平台', inplace=True)
repurchase1 = df_app_user[df_app_user['promote_orders'] == 1]
repurchase_sum = pd.merge(df_serve1, repurchase1, on='user_id', how='inner')
repurchase_sum = repurchase_sum[repurchase_sum['promote_orders_time'] < repurchase_sum['create_time']]
repurchase_sum['create_time'] = repurchase_sum['create_time'].apply(lambda x: x[0:10])
repurchase_sum['promote_orders_time'] = repurchase_sum['promote_orders_time'].apply(lambda x: x[0:10])
repurchase_sum = repurchase_sum[
    repurchase_sum['create_time'].apply(f) != repurchase_sum['promote_orders_time'].apply(f)]

repurchase_sum1 = repurchase_sum.drop_duplicates('user_id')

repurchase_sum1 = repurchase_sum1.groupby('store_id')['user_id'].count().reset_index()
repurchase_sum1 = repurchase_sum1.rename(columns={'user_id': '老客到店人数', 'store_id': 'recommend_store_id'})

df_serve1['create_time'] = df_serve1['create_time'].apply(lambda x: x[0:10])

# ### 新客到店人数

repurchase_sum2 = df_serve1[~df_serve1['create_time'].isin(list(repurchase_sum['create_time']))]

repurchase_sum2.drop_duplicates('user_id', inplace=True)

repurchase_sum2 = repurchase_sum2.groupby('store_id')['user_id'].count().reset_index()
repurchase_sum2 = repurchase_sum2.rename(columns={'user_id': '新客到店人数', 'store_id': 'recommend_store_id'})
repurchase_sum2.sum()

# ### 复购成交人数

repurchase = df_app_user[df_app_user['repurchase_orders'] == 1]
repurchase_people = pd.merge(df_order_remove, repurchase, on='user_id', how='inner')
repurchase_people = repurchase_people.drop_duplicates('user_id')

repurchase_people = repurchase_people.groupby('recommend_store_id')['user_id'].count().reset_index()
repurchase_people = repurchase_people.rename(columns={'user_id': '复购成交人数'})

df_complex1 = pd.merge(df_app_user, df_order_remove, on='user_id', how='inner')  # 合并订单和用户表
df_complex2 = pd.merge(df_app_user, df_recharge, on='user_id', how='inner')  # 合并充值和用户表
df_complex1 = df_complex1[df_complex1['repurchase_orders'] == 1]
df_complex2 = df_complex2[df_complex2['repurchase_orders'] == 1]

# 计算复购订单金额

df_complex1 = df_complex1[df_complex1['repurchase_orders_time'] <= df_complex1['create_date']]

df_complex2 = df_complex2[df_complex2['repurchase_orders_time'] <= df_complex2['recharge_date']]
df_complex2.rename(columns={'amount_pay': 'accounted_money', 'activity_store_id': 'recommend_store_id'}, inplace=True)

df_complex1 = df_complex1.loc[:, ['user_id', 'accounted_money', 'recommend_store_id']]
df_complex2 = df_complex2.loc[:, ['user_id', 'accounted_money', 'recommend_store_id']]

repurchase_money = pd.concat([df_complex2, df_complex1])

# 复购成交金额

repurchase_money = repurchase_money.groupby('recommend_store_id')['accounted_money'].sum().reset_index()
repurchase_money = repurchase_money.rename(columns={'accounted_money': '复购成交金额'})

# ### 商品购买人数,商品订单金额

commodity_order = df_order[df_order['order_type'] == 0]

commodity_order_sum = commodity_order.groupby('recommend_store_id')['accounted_money'].count().reset_index()
commodity_order_sum.rename(columns={'accounted_money': '商品购买人数'}, inplace=True)

commodity_order_money = commodity_order.groupby('recommend_store_id')['accounted_money'].sum().reset_index()
commodity_order_money.rename(columns={'accounted_money': '商品订单金额'}, inplace=True)

# ### 服务购买人数,服务订单金额

service_order = df_order[df_order['order_type'] != 0]

service_order_sum = service_order.groupby('recommend_store_id')['accounted_money'].count().reset_index()
service_order_sum.rename(columns={'accounted_money': '服务购买人数'}, inplace=True)

service_order_money = service_order.groupby('recommend_store_id')['accounted_money'].sum().reset_index()
service_order_money.rename(columns={'accounted_money': '服务订单金额'}, inplace=True)

# ### 充值人数,充值金额,总成交金额,总退款金额

###总充值金额 充值人数

recharge_sum = df_recharge.drop_duplicates('user_id')
recharge_sum = df_recharge.groupby('activity_store_id')['user_id'].count().reset_index()
recharge_sum = recharge_sum.rename(columns={'activity_store_id': 'recommend_store_id', 'user_id': '充值人数'})

recharge_pay = df_recharge.groupby('activity_store_id')['amount_pay'].sum().reset_index()
recharge_pay = recharge_pay.rename(columns={'activity_store_id': 'recommend_store_id', 'amount_pay': '充值金额'})

# 总订单金额

order_pay = df_order.groupby('recommend_store_id')['accounted_money'].sum().reset_index()
order_pay = order_pay.rename(columns={'accounted_money': '总订单金额'})

##### 总成交金额
total_pay = pd.merge(order_pay, recharge_pay, on='recommend_store_id', how='outer')
total_pay.fillna(0, inplace=True)
total_pay.eval(""" 总成交金额= 总订单金额 + 充值金额 """, inplace=True)
total_pay = total_pay.loc[:, ['recommend_store_id', '总成交金额']]

# 总成交人数

total_sum = df_order.drop_duplicates('user_id')

total_sum = total_sum.groupby('recommend_store_id')['user_id'].count().reset_index()
total_sum = total_sum.rename(columns={'user_id': '总成交人数'})

# ### 自拓客

plan = conn_epmall['epmall_deliver_plan_100063']  # 客户渠道表
plan = pd.DataFrame(list(plan.find({}, {'_id': 0, 'deliver_page.link_id': 1, 'deliver_id': 1})))
plan['deliver_page'] = plan['deliver_page'].apply(lambda x: str(x)[13:-2])

# 自拓客到店人数
plan_sum = pd.merge(df_serve, df_app_user, how='left')
plan_sum = pd.merge(plan_sum, plan, on='deliver_id', how='left')
plan_sum = plan_sum[plan_sum['store_id'] == plan_sum['deliver_page']]
plan_sum.drop_duplicates('user_id', inplace=True)
plan_sum = plan_sum.groupby('store_id')['user_id'].count().reset_index()
plan_sum = plan_sum.rename(columns={'user_id': '自拓客到店人数', 'store_id': 'recommend_store_id'})

###自拓客实收金额
# 自拓客实收订单金额
plan_money = pd.merge(df_order, df_app_user, how='left')
plan_money = pd.merge(plan_money, plan, on='deliver_id', how='left')
plan_money = plan_money[plan_money['recommend_store_id'] == plan_money['deliver_page']]

# 自拓客实收充值金额
plan_money1 = pd.merge(df_recharge, df_app_user, how='left')
plan_money1 = pd.merge(plan_money1, plan, on='deliver_id', how='left')
plan_money1 = plan_money1[plan_money1['activity_store_id'] == plan_money1['deliver_page']]

plan_money1.rename(columns={'amount_pay': 'accounted_money', 'activity_store_id': 'recommend_store_id'}, inplace=True)
plan_money = plan_money.loc[:, ['user_id', 'accounted_money', 'recommend_store_id']]
plan_money1 = plan_money1.loc[:, ['user_id', 'accounted_money', 'recommend_store_id']]

# 自拓客实收金额=自拓客实收充值金额+自拓客实收订单金额
plan_money2 = pd.concat([plan_money, plan_money1])
plan_money2 = plan_money2.groupby('recommend_store_id')['accounted_money'].sum().reset_index()
plan_money2.rename(columns={'accounted_money': '自拓客实收金额'}, inplace=True)

# 自拓客成交人数
plan_money_sum = pd.merge(df_order, df_app_user, how='left')
plan_money_sum = pd.merge(plan_money_sum, plan, on='deliver_id', how='left')
plan_money_sum = plan_money_sum[plan_money_sum['recommend_store_id'] == plan_money_sum['deliver_page']]
plan_money_sum.drop_duplicates('user_id', inplace=True)
plan_money_sum = plan_money_sum.groupby('recommend_store_id')['user_id'].count().reset_index()
plan_money_sum.rename(columns={'user_id': '自拓客成交人数'}, inplace=True)

# ### 总退款金额
df_back_new = None
df_refund_new = None

# #充值退款
if len(df_refund) != 0:
    df_refund_new = pd.merge(df_refund, df_recharge_all, on='deal_no', how='left')

    df_refund_new = df_refund_new.groupby('activity_store_id')['amount_refund'].sum().reset_index()
    df_refund_new = df_refund_new.rename(columns={'activity_store_id': 'recommend_store_id'})

if len(df_back) != 0:
    df_back_new = pd.merge(df_back, df_order_all0, on='order_no', how='left')  # 订单退款

    df_back_new = df_back_new.loc[:, ['order_no', 'recommend_store_id', 'wxrefund_money']]
    df_back_new = df_back_new.groupby('recommend_store_id')['wxrefund_money'].sum().reset_index()

total_money = None
if len(df_back) != 0 and len(df_refund) != 0:
    total_money = pd.merge(df_back_new, df_refund_new, on='recommend_store_id', how='outer')
    total_money.fillna(0, inplace=True)
    total_money.eval(""" 总退款金额= wxrefund_money + amount_refund """, inplace=True)
    total_money = total_money.loc[:, ['recommend_store_id', '总退款金额']]

if len(df_back) == 0 and len(df_refund) != 0:
    total_money = df_refund_new.rename(columns={'amount_refund': '总退款金额'})
if len(df_refund) == 0 and len(df_back) != 0:
    total_money = df_back_new.rename(columns={'wxrefund_money': '总退款金额'})

# 体验卡退款commodity_id_list

df_back_yes = None
if len(df_back) != 0:
    df_back_yes = df_back[df_back['commodity_id'].isin(commodity_id_list)]  # 筛选体验卡退款
    df_back_yes = pd.merge(df_back_yes, df_order_all0, on='order_no', how='left')  # 订单退款

    df_back_yes = df_back_yes.loc[:, ['order_no', 'recommend_store_id', 'wxrefund_money']]
    df_back_yes = df_back_yes.groupby('recommend_store_id').agg(
        {'order_no': 'count', 'wxrefund_money': 'sum'}).reset_index()
    df_back_yes = df_back_yes.rename(columns={'wxrefund_money': '体验卡退款金额', 'order_no': '体验卡退款数量'})

# ### 合并

if len(df_refund) != 0 or len(df_back) != 0:
    dfs = [total_sum, total_pay, total_money, recharge_sum, recharge_pay, repurchase_money, repurchase_people,
           repurchase_sum1, repurchase_sum2, card_money, card_sum, up_money, up_sum, commodity_order_sum,
           commodity_order_money, service_order_sum, service_order_money, serve_sum, serve_money, plan_sum,
           plan_money_sum, plan_money2]
else:
    dfs = [total_sum, total_pay, recharge_sum, recharge_pay, repurchase_money, repurchase_people, repurchase_sum1,
           repurchase_sum2, card_money, card_sum, up_money, up_sum, commodity_order_sum, commodity_order_money,
           service_order_sum, service_order_money, serve_sum, serve_money, plan_sum, plan_money_sum, plan_money2]

if df_back_yes is not None:
    dfs.append(df_back_yes)

df_result = reduce(lambda left, right: pd.merge(left, right, on='recommend_store_id', how='outer'), dfs)
platform = df_result[df_result['recommend_store_id'] == '平台']
platform['store_name'] = '平台'

df_store = conn_epmall['epmall_store_100063']  # 门店
df_store = pd.DataFrame(list(df_store.find({}, {'_id': 0, 'store_id': 1, 'store_name': 1, 'city': 1})))
df_result = pd.merge(df_result, df_store, left_on='recommend_store_id', right_on='store_id', how='inner')
df_result.drop('store_id', axis=1, inplace=True)
df_result = pd.concat([platform, df_result])
df_result = df_result.rename(columns={'city': '城市'})
# In[268]:


df_result.fillna(0, inplace=True)

if len(df_refund) != 0 or len(df_back) != 0:
    df_result.eval("""
    退款率 = 总退款金额 / 总成交金额
    售卡金额占比 = 售卡金额 / 总成交金额
    复购率 = 复购成交人数/老客到店人数
    复购金额占比 = 复购成交金额/总成交金额
    实际成交金额 = 总成交金额-总退款金额
    充值占比 = 充值金额/总成交金额
    商品订单金额占比 = 商品订单金额/总成交金额
    服务订单金额占比 = 服务订单金额/总成交金额
    升单率 = 升单人数/新客到店人数
    升单金额占比 = 升单金额/总成交金额 """, inplace=True)
else:
    df_result.eval("""
        退款率 = 0
        总退款金额=0
        售卡金额占比 = 售卡金额 / 总成交金额
        复购率 = 复购成交人数/老客到店人数
        复购金额占比 = 复购成交金额/总成交金额
        实际成交金额 = 总成交金额-总退款金额
        充值占比 = 充值金额/总成交金额
        商品订单金额占比 = 商品订单金额/总成交金额
        服务订单金额占比 = 服务订单金额/总成交金额
        升单率 = 升单人数/新客到店人数
        升单金额占比 = 升单金额/总成交金额 """, inplace=True)

df_result['store_name'] = df_result['store_name'].apply(lambda x: x.strip())

db = pymysql.connect(host='l5vhlstu.ipyingshe.net', database='sys', user='root', password='root123', port=31453,
                     charset='utf8')
sql = 'select * from Sheet2'
df_region = pd.read_sql(sql, db)

df_result = pd.merge(df_region, df_result, left_on='门店ID', right_on='recommend_store_id', how='right')
df_result.info()

if len(df_back) == 0:
    df_result['体验卡退款金额'] = 0
    df_result['体验卡退款数量'] = 0
df_result = df_result.loc[:,
            ['大区', '门店名称', '城市', '总成交人数', '总成交金额', '总退款金额', '体验卡退款金额', '体验卡退款数量', '实际成交金额', '退款率', '充值人数', '充值金额',
             '充值占比', '商品购买人数',
             '商品订单金额', '商品订单金额占比',
             '服务购买人数', '服务订单金额', '服务订单金额占比', '售卡人数', '售卡金额', '售卡金额占比', '新客到店人数', '升单人数', '升单金额', '升单金额占比', '升单率',
             '老客到店人数', '复购成交人数', '复购成交金额', '复购金额占比', '复购率', '服务人数', '服务金额', '自拓客到店人数', '自拓客成交人数',
             '自拓客实收金额']]

df_result.sort_values("大区", inplace=True)

df_advertise = df_result.loc[:,
               ['城市', '新客到店人数', '老客到店人数', '售卡人数', '售卡金额', '升单人数', '升单金额', '复购成交人数', '复购成交金额', '实际成交金额', '总退款金额',
                '总成交金额']]

df_advertise = df_advertise.groupby(['城市']).sum().reset_index()

df_advertise['城市'] = df_advertise['城市'].mask(df_advertise['城市'] == 0, '平台')

df_advertise['总到店人数'] = df_advertise['新客到店人数'] + df_advertise['老客到店人数']
df_advertise['升单率'] = df_advertise['升单人数'] / df_advertise['新客到店人数']
df_advertise['复购率'] = df_advertise['复购成交人数'] / df_advertise['老客到店人数']
df_advertise['客单价'] = (df_advertise['升单金额'] + df_advertise['复购成交金额']) / (df_advertise['升单人数'] + df_advertise['复购成交人数'])
df_advertise['退款金额占比'] = df_advertise['总退款金额'] / df_advertise['实际成交金额']
list_advertise = ['复购率', '退款金额占比', '升单率']

df_advertise['城市'] = df_advertise['城市'].apply(lambda x: x.replace('市', ''))

df_temp_table = pd.merge(df_middle_result, df_advertise, on='城市', how='left')

df_temp_table['有效进线人数'] = df_temp_table['进线人数'] - df_temp_table['重粉数']

df_temp_table['有效进线成本'] = df_temp_table['广告总额'] / df_temp_table['有效进线人数']

df_temp_table['到店成本'] = df_temp_table['广告总额'] / df_temp_table['总到店人数']

df_temp_table['投产比'] = df_temp_table['实际成交金额'] / df_temp_table['广告总额']

df_temp_table = df_temp_table.loc[:,
                ['负责人', '大区', '省份', '城市', '广告总额', '进线成本', '进线人数', '系统重粉数', '表单重粉数', '重粉数', '重粉率', '有效进线成本', '到店成本',
                 '总到店人数', '新客到店人数',
                 '老客到店人数', '售卡人数', '售卡金额', '升单人数', '升单金额', '升单率', '复购成交人数', '复购成交金额', '复购率', '客单价', '实际成交金额',
                 '总退款金额', '总成交金额', '退款金额占比', '投产比']]

df_temp_table = df_temp_table.fillna(0)
df_temp_table = df_temp_table.replace([np.inf, -np.inf], 0)

df_temp_table.to_excel('区域广告投放.xlsx', index=False)

date1 = str(yesterday)[5:]
date2 = date1.replace('-', '.')

from openpyxl import load_workbook

wb = load_workbook(filename="区域广告投放.xlsx")
ws = wb.active

for col in ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in col:
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(name='黑体', size=11)

prj = ws.columns
people_list = []
prjTuple = tuple(prj)
for idx in range(0, 1):  # 遍历A、B、C、D列  #(0,len(prjTuple)) 遍历全部列
    for cell in prjTuple[idx]:
        if (cell.value != '负责人' and cell.value != '平台' and cell.value is not None):
            #         print(cell.coordinate, cell.value)
            people_list.append(cell.value)

people_dict = {}
for i in people_list:
    people_dict[i] = people_list.count(i)

people_list1 = []
for j in people_list:
    if j not in people_list1:
        people_list1.append(j)

font0 = Font(name='黑体', size=11)
font1 = Font(name='黑体', size=11, b=True)  # 字体
font2 = Font(name='黑体', size=16, b=True)
line_t = Side(style='thin', color='000000')  # 细边框
border1 = Border(top=line_t, bottom=line_t, left=line_t, right=line_t)
alignment = Alignment(horizontal='center', vertical='center')  # 居中

fill1 = PatternFill('solid', fgColor='FFCC33')  # 浅黄色
fill2 = PatternFill('solid', fgColor='6495ED')  # 蓝色

sty0 = NamedStyle(name='sty0', font=font0, border=border1, alignment=alignment)
sty1 = NamedStyle(name='sty1', font=font1, border=border1, alignment=alignment, fill=fill1)
sty2 = NamedStyle(name='sty2', font=font1, border=border1, alignment=alignment, fill=fill2)
sty3 = NamedStyle(name='sty3', font=font2, border=border1, alignment=alignment)


def get_cal(row_n):
    wb.guess_types = True
    # 进线成本
    adv_total = ws.cell(row=row_n, column=5).value  # 广告总额
    inline_people = ws.cell(row=row_n, column=7).value  # 进线人数
    if inline_people == 0:
        inline_cost = 0
    else:
        inline_cost = np.divide(adv_total, inline_people)
    ws.cell(row=row_n, column=6, value=inline_cost)
    # 重粉率
    dup_people = ws.cell(row=row_n, column=10).value  # 重粉数
    if inline_people == 0:
        dup_rate = '0%'
    else:
        dup_rate = np.divide(dup_people, inline_people)
    ws.cell(row=row_n, column=11, value=dup_rate)
    # 有效进线成本
    valid_people = inline_people - dup_people
    if valid_people == 0:
        valid_inline_cost = 0
    else:
        valid_inline_cost = np.divide(adv_total, valid_people)
    ws.cell(row=row_n, column=12, value=valid_inline_cost)
    # 到店成本
    total_arrivals = ws.cell(row=row_n, column=14).value  # 总到店人数
    if total_arrivals == 0:
        arrivals_cost = 0
    else:
        arrivals_cost = np.divide(adv_total, total_arrivals)
    ws.cell(row=row_n, column=13, value=arrivals_cost)
    # 升单率计算  升单率 = 升单人数/新客到店人数
    upgrade_order_num = ws.cell(row=row_n, column=19).value
    new_people_num = ws.cell(row=row_n, column=15).value
    if new_people_num == 0:
        upgrade_order_rate = 0
    else:
        upgrade_order_rate = np.divide(upgrade_order_num, new_people_num)
    ws.cell(row=row_n, column=21, value=upgrade_order_rate)
    # 复购率  复购率 = 复购成交人数/老客到店人数
    repurchase_done = ws.cell(row=row_n, column=22).value
    frequent_custormer = ws.cell(row=row_n, column=16).value
    if frequent_custormer == 0:
        repurchase_rate = 0
    else:
        repurchase_rate = np.divide(repurchase_done, frequent_custormer)
    ws.cell(row=row_n, column=24, value=repurchase_rate)
    # 客单价
    upgrade_amount = ws.cell(row=row_n, column=20).value  # 升单金额
    repurchase_amount = ws.cell(row=row_n, column=23).value  # 复购金额
    if upgrade_order_num + repurchase_done == 0:
        per_ticket_sales = 0
    else:
        per_ticket_sales = np.divide(upgrade_amount + repurchase_amount, upgrade_order_num + repurchase_done)
    ws.cell(row=row_n, column=25, value=per_ticket_sales)
    # 退款金额占比
    total_refund = ws.cell(row=row_n, column=27).value
    actually_done_amount = ws.cell(row=row_n, column=26).value  # 实际成交金额
    refund_ptoprotion = np.divide(total_refund, actually_done_amount)
    ws.cell(row=row_n, column=29, value=refund_ptoprotion)
    # 投产比
    if adv_total == 0:
        input_output_ratio = '1/0.00'
    else:
        input_output_ratio = np.divide(actually_done_amount, adv_total)
    ws.cell(row=row_n, column=30, value=input_output_ratio)


# 合计计算
sum_list = [5, 7, 8, 9, 10, 14, 15, 16, 17, 18, 19, 20, 22, 23, 26, 27, 28]
last_row = ws.max_row + 1
for s_column in sum_list:
    total = 0
    for f in range(2, last_row):
        one_value = ws.cell(row=f, column=s_column).value
        total += one_value
    ws.cell(row=last_row, column=s_column, value=total)
get_cal(last_row)

result = 1
for k in range(len(people_list1)):
    result = result + people_dict[people_list1[k]] + 1
    # 插入一行
    ws.insert_rows(idx=result, amount=1)
    # 计算值
    start_index = result - people_dict[people_list1[k]]
    end_index = result
    for sum_column in sum_list:
        res = 0
        for index in range(start_index, end_index):
            cellValue = ws.cell(row=index, column=sum_column).value
            res += cellValue
        ws.cell(row=result, column=sum_column, value=res)
    get_cal(result)

ws.insert_rows(idx=0, amount=1)
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=30)
ws.cell(1, 1).value = f'区 域 广 告 投 放 转 化 表（{date2}）'
ws.cell(1, 1).style = sty3

res1 = 2
for s in range(len(people_list1)):
    res1 = res1 + people_dict[people_list1[s]] + 1
    # 合并合计单元格
    ws.cell(res1, 1).value = '汇总'
    ws.cell(res1, 1).style = sty1
    for x in range(2, 31):
        ws.cell(res1, x).style = sty1
    start_area = res1 - people_dict[people_list1[s]]
    ws.merge_cells(start_row=start_area, start_column=1, end_row=res1 - 1, end_column=1)
    ws.cell(start_area, 1).value = people_list1[s]
    ws.cell(start_area, 1).style = sty0
final_row = ws.max_row
ws.cell(final_row, 1).value = '合计'
ws.cell(final_row, 1).style = sty1
for y in range(2, 31):
    ws.cell(final_row, y).style = sty1

# 第一行表头
for m in range(1, 31):
    ws.cell(2, m).style = sty2
ws.row_dimensions[2].height = 24  # 行高

list_percentage = ['K', 'U', 'X', 'AC']
for x in list_percentage:
    for y in range(2, ws.max_row + 1):
        cell_postition = x + str(y)
        ws[cell_postition].number_format = '0%'  # 设置成百分比格式

list_percentage1 = ['F', 'L', 'M', 'Y']
for m in list_percentage1:
    for n in range(2, ws.max_row + 1):
        cell_postition1 = m + str(n)
        ws[cell_postition1].number_format = '0'  # 设置成整数格式

for n in range(2, ws.max_row + 1):
    cell_postition3 = 'E' + str(n)
    ws[cell_postition3].number_format = '0.00'  # 设置成两位小数格式

for w in range(2, ws.max_row + 1):
    cell_postition2 = 'AD' + str(w)
    ws[cell_postition2].number_format = '1!/0.00'  # 设置成分数格式

# 行高
ws.row_dimensions[1].height = 24
# 调整列宽
ws.column_dimensions['E'].width = 13.13


# 边框设置
def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


s = 'A1:AD' + str(ws.max_row)
set_border(ws, s)

wb.save(filename=f"区域广告投放转化表({date2}).xlsx")


def upload_file(file_path, wx_upload_url):
    file_name = file_path.split("/")[-1]
    with open(file_path, 'rb') as f:
        length = os.path.getsize(file_path)
        data = f.read()
    headers = {"Content-Type": "application/octet-stream"}
    params = {
        "filename": file_name,
        "filelength": length,
    }
    file_data = copy.copy(params)
    file_data['file'] = (file_path.split('/')[-1:][0], data)
    encode_data = encode_multipart_formdata(file_data)
    file_data = encode_data[0]
    headers['Content-Type'] = encode_data[1]
    r = requests.post(wx_upload_url, data=file_data, headers=headers)
    media_id = r.json()['media_id']
    return media_id


def qi_ye_wei_xin_file(wx_url, media_id):
    headers = {"Content-Type": "text/plain"}
    data = {
        "msgtype": "file",
        "file": {
            "media_id": media_id
        }
    }
    r = requests.post(
        url=wx_url,
        headers=headers, json=data)


test_report = f'区域广告投放转化表({date2}).xlsx'

wx_api_key = '457aa3c8-201c-4583-9f6f-2747286bed2c'
wx_upload_url = "https://qyapi.weixin.qq.com/cgi-bin/webhook/upload_media?key={}&type=file".format(wx_api_key)
wx_url = 'https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key={}'.format(wx_api_key)
media_id = upload_file(test_report, wx_upload_url)
qi_ye_wei_xin_file(wx_url, media_id)
os.remove(f"区域广告投放转化表({date2}).xlsx")

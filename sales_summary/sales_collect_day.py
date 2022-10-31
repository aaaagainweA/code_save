#!/usr/bin/env python
# coding: utf-8

import pandas as pd
import excel2img
import datetime
import requests
import base64
import hashlib
import os
from copy import copy
from urllib3 import encode_multipart_formdata

today = datetime.date.today()
oneday = datetime.timedelta(days=1)
yesterday = today - oneday
date=str(yesterday)
date1=str(today)

#-----------销售日报计算-----------

df1=pd.read_excel(r'D:/python_project/projects/region_adverting/city_consume.xlsx')

#当日投放
throwin_yesterday=df1['广告总额'].sum()
#当日总进线
inline_yesterday=df1['进线人数'].sum()
#单粉成本
single_cost=throwin_yesterday/inline_yesterday

df2=pd.read_excel(r'D:/python_project/projects/national_sales_code/全国市场门店销售日报表临时用.xlsx')

list_ccc=['平台','广东江苏河南区域','武汉-北京','浙江区域','西南区域']

df2=df2.loc[df2['大区'].isin(list_ccc),:]
#复购金额
repuchase_amount=df2['复购成交金额'].sum()

#商品单业绩 
product_order_amount=df2['商品订单金额'].sum()
#服务单业绩
serve_order_amount=df2['服务订单金额'].sum()
#充值金额
recharge_amount=df2['充值金额'].sum()
#退款总额
refund_amount=df2['总退款金额'].sum()

#当日合计销售金额
actual_transaction_amount=df2['实际成交金额'].sum()
#投产比
input_output_ratio=actual_transaction_amount/throwin_yesterday

#线下成交人数
upgrade_people=df2['升单人数'].sum()
repuchase_people=df2['复购成交人数'].sum()
offline_deal_people=upgrade_people+repuchase_people

#退卡数
card_return_num=df2['体验卡退款数量'].sum()
#退卡金额
card_return_amount=df2['体验卡退款金额'].sum()

df3=pd.read_excel('礼宾客服.xlsx')

df3=df3.loc[df3['实付金额']>0,:]
df5=df3.copy()

df3.astype({'订单创建日期':'str'})
df3=df3.loc[df3['订单创建日期'].str.contains(date),:]

#千元以上卡项数量
over_thousand_num=df3.loc[(df3['实付金额']>=999)&(df3['套餐分类']=='礼宾售卡')&(df3['是否退款']=='否'),'订单编号'].count()
#千元以下卡项数量
loewr_thousand_num=df3.loc[(df3['实付金额']<999)&(df3['套餐分类']=='礼宾售卡')&(df3['是否退款']=='否'),'订单编号'].count()
#售卡数量合计
card_total=df3.loc[(df3['套餐分类']=='礼宾售卡')&(df3['是否退款']=='否'),'订单编号'].count()
#千元以上卡占比
over_thousand_percent=over_thousand_num/card_total

#升级卡金额
upgrade_card_amount=df3.loc[df3['套餐分类']=='礼宾售卡升单','实付金额'].sum()
##线上售卡金额合计
card_amount=df3.loc[df3['是否退款']=='否','实付金额'].sum()-upgrade_card_amount
#线下成交金额合计
offline_amount_total=actual_transaction_amount-card_amount+refund_amount
#线下成交客单价
offline_per_sales=(offline_amount_total-upgrade_card_amount)/offline_deal_people
#退款金额
return_amount=refund_amount-card_return_amount


#升单金额
upgrade_order_amount=offline_amount_total-upgrade_card_amount-repuchase_amount


#-----------销售月报计算-----------

df4=pd.read_excel(r'D:/python_project/projects/national_sales_code/全国市场门店销售月报表临时用.xlsx')

df4=df4.loc[df4['大区'].isin(list_ccc),:]

#复购金额
repuchase_amounts=df4['复购成交金额'].sum()

#商品单业绩 
product_order_amounts=df4['商品订单金额'].sum()
#服务单业绩
serve_order_amounts=df4['服务订单金额'].sum()
#充值金额
recharge_amounts=df4['充值金额'].sum()
#退款总额
refund_amounts=df4['总退款金额'].sum()

#当日合计销售金额
actual_transaction_amounts=df4['实际成交金额'].sum()

#线下成交人数
upgrade_people_all=df4['升单人数'].sum()
repuchase_people_all=df4['复购成交人数'].sum()
offline_deal_number=upgrade_people_all+repuchase_people_all

#退卡数
card_return_number=df4['体验卡退款数量'].sum()
#退卡金额
card_return_amounts=df4['体验卡退款金额'].sum()

df5.astype({'订单创建日期':'str'})
df5=df5.loc[~(df5['订单创建日期'].str.contains(date1)),:]

#千元以上卡项数量
over_thousand_number=df5.loc[(df5['实付金额']>=999)&(df5['套餐分类']=='礼宾售卡')&(df5['是否退款']=='否'),'订单编号'].count()
#千元以下卡项数量
loewr_thousand_number=df5.loc[(df5['实付金额']<999)&(df5['套餐分类']=='礼宾售卡')&(df5['是否退款']=='否'),'订单编号'].count()
#售卡数量合计
card_totals=df5.loc[(df5['套餐分类']=='礼宾售卡')&(df5['是否退款']=='否'),'订单编号'].count()
#千元以上卡占比
over_thousand_percents=over_thousand_number/card_totals

#升级卡金额
upgrade_card_amounts=df5.loc[df5['套餐分类']=='礼宾售卡升单','实付金额'].sum()
#线上售卡金额合计
card_amounts=df5.loc[df5['是否退款']=='否','实付金额'].sum()-upgrade_card_amounts
#线下成交金额合计
offline_amount_totals=actual_transaction_amounts-card_amounts+refund_amounts
#线下成交客单价
offline_per_sale=(offline_amount_totals-upgrade_card_amounts)/offline_deal_number
#退款金额
return_amounts=refund_amounts-card_return_amounts


#升单金额
upgrade_order_amounts=offline_amount_totals-repuchase_amounts-upgrade_card_amounts


df6=pd.read_excel(r'D:/python_project/projects/national_sales_code/各城市广告费累计.xlsx')


#当月投放
put_in_month=df6['广告总额'].sum()

#当月总进线
total_inline_month=df6['进线人数'].sum()



#单粉成本
single_cost_month=put_in_month/total_inline_month


#整体投产比
in_out_ratio=actual_transaction_amounts/put_in_month



from openpyxl import load_workbook
wb = load_workbook(filename="销售汇总表模板.xlsx")
ws = wb.active



#日表写入



ws.cell(row=3, column=1, value=throwin_yesterday)#当日投放
ws.cell(row=3, column=2, value=inline_yesterday)#当日总进线
ws.cell(row=3, column=5, value=single_cost)#单粉成本
ws.cell(row=3, column=9, value=card_amount)#线上售卡金额合计
ws.cell(row=5, column=7, value=upgrade_order_amount)#升单金额
ws.cell(row=5, column=8, value=repuchase_amount)#复购金额
ws.cell(row=7, column=6, value=product_order_amount)#商品单业绩 
ws.cell(row=7, column=8, value=serve_order_amount)#服务单业绩
ws.cell(row=7, column=10, value=recharge_amount)#充值金额
ws.cell(row=7, column=12, value=refund_amount)#退款总额
ws.cell(row=8, column=9, value=actual_transaction_amount)#当日合计销售金额
ws.cell(row=8, column=11, value=input_output_ratio)#投产比
ws.cell(row=5, column=9, value=offline_deal_people)#线下成交人数



ws.cell(row=3, column=6, value=over_thousand_num)#千元以上卡项数量
ws.cell(row=3, column=7, value=loewr_thousand_num)#千元以下卡项数量
ws.cell(row=3, column=8, value=card_total)#售卡数量合计
ws.cell(row=3, column=10, value=over_thousand_percent)#千元以上卡占比
ws.cell(row=3, column=11, value=card_return_num)#退卡数
ws.cell(row=3, column=12, value=card_return_amount)#退卡金额
ws.cell(row=5, column=6, value=upgrade_card_amount)#升级卡金额
ws.cell(row=5, column=10, value=offline_amount_total)#线下成交金额合计
ws.cell(row=5, column=11, value=offline_per_sales)#线下成交客单价
ws.cell(row=5, column=12, value=return_amount)#退款金额


#月表写入


ws.cell(row=3, column=21, value=card_amounts)#线上售卡金额合计
ws.cell(row=5, column=19, value=upgrade_order_amounts)#升单金额
ws.cell(row=5, column=20, value=repuchase_amounts)#复购金额
ws.cell(row=7, column=18, value=product_order_amounts)#商品单业绩 
ws.cell(row=7, column=20, value=serve_order_amounts)#服务单业绩
ws.cell(row=7, column=22, value=recharge_amounts)#充值金额
ws.cell(row=7, column=24, value=refund_amounts)#退款总额
ws.cell(row=8, column=21, value=actual_transaction_amounts)#当日合计销售金额
ws.cell(row=5, column=21, value=offline_deal_number)#线下成交人数


ws.cell(row=3, column=18, value=over_thousand_number)#千元以上卡项数量
ws.cell(row=3, column=19, value=loewr_thousand_number)#千元以下卡项数量
ws.cell(row=3, column=20, value=card_totals)#售卡数量合计
ws.cell(row=3, column=22, value=over_thousand_percents)#千元以上卡占比
ws.cell(row=3, column=23, value=card_return_number)#退卡数
ws.cell(row=3, column=24, value=card_return_amounts)#退卡金额
ws.cell(row=5, column=18, value=upgrade_card_amounts)#升级卡金额
ws.cell(row=5, column=22, value=offline_amount_totals)#线下成交金额合计
ws.cell(row=5, column=23, value=offline_per_sale)#线下成交客单价
ws.cell(row=5, column=24, value=return_amounts)#退款金额
ws.cell(row=3, column=13, value=put_in_month)#当月投放
ws.cell(row=3, column=14, value=total_inline_month)#当月总进线
ws.cell(row=3, column=17, value=single_cost_month)#单粉成本
ws.cell(row=8, column=23, value=in_out_ratio)#整体投产比

ws.cell(1, 1).value = f'医普茂销售日报（0点-0点）{date}'

wb.save(filename = f'医普茂当日销售汇总表{date}.xlsx')

excel2img.export_img(f'医普茂当日销售汇总表{date}.xlsx', '医普茂当日销售汇总表'+date+'.png', "Sheet1", None)

def wx_image(image):
    with open(image, 'rb') as file:  # 转换图片成base64格式
        data = file.read()
        encodestr = base64.b64encode(data)
        image_data = str(encodestr, 'utf-8')

    with open(image, 'rb') as file:  # 图片的MD5值
        md = hashlib.md5()
        md.update(file.read())
        image_md5 = md.hexdigest()

    url = "https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=b3ab38ac-adbc-489d-adc1-dadd61c2f98c"  # 填上机器人Webhook地址
    headers = {"Content-Type": "application/json"}
    data = {
        "msgtype": "image",
        "image": {
            "base64": image_data,
            "md5": image_md5
        }
    }
    result = requests.post(url, headers=headers, json=data)
    return result

wx_image(f'当日销售汇总表{date}.png')  # 传入图片路径

def upload_file(file_path, wx_upload_url):
    file_name = file_path.split("/")[-1]
    with open(file_path, 'rb') as f:
        length = os.path.getsize(file_path)
        data = f.read()
    headers = {"Content-Type": "application/octet-stream"}
    params = {
        "filename": file_name,
        "filelength": length,
    }
    file_data = copy(params)
    file_data['file'] = (file_path.split('/')[-1:][0], data)
    encode_data = encode_multipart_formdata(file_data)
    file_data = encode_data[0]
    headers['Content-Type'] = encode_data[1]
    r = requests.post(wx_upload_url, data=file_data, headers=headers)
    media_id = r.json()['media_id']
    return media_id

def qi_ye_wei_xin_file(wx_url, media_id):
    headers = {"Content-Type": "text/plain"}
    data = {
        "msgtype": "file",
        "file": {
            "media_id": media_id
        }
    }
    r = requests.post(
        url=wx_url,
        headers=headers, json=data)

test_report = f'当日销售汇总表{date}.xlsx'

wx_api_key = 'b3ab38ac-adbc-489d-adc1-dadd61c2f98c'
wx_upload_url = "https://qyapi.weixin.qq.com/cgi-bin/webhook/upload_media?key={}&type=file".format(wx_api_key)
wx_url = 'https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key={}'.format(wx_api_key)
media_id = upload_file(test_report, wx_upload_url)
qi_ye_wei_xin_file(wx_url, media_id)

os.remove(f'当日销售汇总表{date}.png')
os.remove(f'当日销售汇总表{date}.xlsx')
